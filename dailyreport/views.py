import csv, logging, openpyxl
from io import BytesIO
from datetime import datetime, date, timedelta, time as dtime, time
from tempfile import NamedTemporaryFile
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from calendar import monthrange
from django.conf import settings

# ===== BEGIN TEMPLATE_VERSION CONST M0 =====
# 外部日報 Excel テンプレートの期待バージョン
TEMPLATE_VERSION = "2025.01"

# 校验用：Excel 里读出来的 version_val 必须等于这个值
EXPECTED_TEMPLATE_VERSION = TEMPLATE_VERSION
# ===== END INSERT =====

from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import render, get_object_or_404, redirect, HttpResponse
from django.utils.timezone import now, make_aware, get_current_timezone
TZ = get_current_timezone()
from django.utils import timezone
from django.db import transaction
from django.db.models import IntegerField, Value, Case, When, ExpressionWrapper, F, Sum, Q, DateField
from django.db.models.functions import Substr, Cast, Coalesce, NullIf, Lower, Trim
from django.http import HttpResponse, FileResponse
from django.utils.encoding import escape_uri_path
from django.urls import reverse
from django.utils.http import urlencode
from django.forms import inlineformset_factory
from dateutil.relativedelta import relativedelta

from dailyreport.constants import PAYMENT_RATES, CHARTER_CASH_KEYS, CHARTER_UNCOLLECTED_KEYS
from dailyreport.models import DriverDailyReport, DriverDailyReportItem
from .forms import DriverDailyReportForm, DriverDailyReportItemForm, ReportItemFormSet, RequiredReportItemFormSet, ExternalDailyReportImportForm
from .services.calculations import calculate_deposit_difference
from dailyreport.services.summary import (
    resolve_payment_method,
    calculate_totals_from_instances, calculate_totals_from_formset
)
from dailyreport.utils.debug import debug_print



from staffbook.services import get_driver_info
from staffbook.models import Driver
from carinfo.models import Car

from vehicles.models import Reservation
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
# ===== BEGIN INSERT DUP-IMPORT-1 =====
from io import BytesIO
import base64
# ===== END INSERT DUP-IMPORT-1 =====
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def _norm_hhmm(v: object) -> str:
    """
    归一化为 'HH:MM'；不合法返回 ''。
    支持：
      - datetime/time 对象
      - '10:30'、'10：30'（全角冒号）
      - '1030' 或 '930'
    """
    if not v:
        return ""
    if isinstance(v, dtime):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%H:%M")

    s = str(v).strip().replace("：", ":")
    if not s:
        return ""

    # 纯数字 3~4 位：930 / 1030
    if s.isdigit() and len(s) in (3, 4):
        h = int(s[:-2]); m = int(s[-2:])
        if 0 <= h < 24 and 0 <= m < 60:
            return f"{h:02d}:{m:02d}"
        return ""

    if ":" in s:
        try:
            h, m = map(int, s.split(":", 1))
            if 0 <= h < 24 and 0 <= m < 60:
                return f"{h:02d}:{m:02d}"
        except Exception:
            return ""
    return ""

def _as_aware_dt(val, base_date):
    """把 datetime / time / 'HH:MM' 统一成当天的 aware datetime。"""
    if not val:
        return None
    if isinstance(val, datetime):
        return val if val.tzinfo else make_aware(val, TZ)
    if isinstance(val, dtime):
        return make_aware(datetime.combine(base_date, val), TZ)
    s = str(val).strip()
    if ":" in s:
        try:
            h, m = map(int, s.split(":", 1))
            return make_aware(datetime.combine(base_date, dtime(h, m)), TZ)
        except Exception:
            return None
    return None
# =================================================

# === Deposit summary helper (统一口径：ながし現金 + 貸切現金) ===
def _build_deposit_summary_from_totals_raw(totals_raw: dict, deposit_amount: int | Decimal | None):
    """
    totals_raw: 来自 calculate_totals_from_formset / calculate_totals_from_instances 的原始结构
      - 貸切現金: totals_raw["charter_cash_total"]
      - ながし現金: totals_raw["nagashi_cash"]["total"]
    """
    charter_cash = int(totals_raw.get("charter_cash_total", 0) or 0)
    nagashi_cash = int((totals_raw.get("nagashi_cash") or {}).get("total", 0) or 0)
    expected = charter_cash + nagashi_cash
    deposit = int(deposit_amount or 0)
    return {
        "expected_deposit": expected,         # 应入金 = ながし現金 + 貸切現金
        "deposit_amount": deposit,            # 实入金（表头的入金）
        "deposit_difference": deposit - expected,  # 差额 = 实入金 - 应入金
    }


# === Uber 别名 & 关键词：统一口径（导出/总览共用） ===
import re as _re

TIP_PAT   = _re.compile(r'(チップ|tip|小费|ﾁｯﾌﾟ)', _re.IGNORECASE)
RESV_PAT  = _re.compile(r'(予約|reservation)', _re.IGNORECASE)
PROMO_PAT = _re.compile(r'(プロモ|promotion)', _re.IGNORECASE)

UBER_TIP_ALIASES   = {'uber_tip', 'uber tip', 'ubertip'}
UBER_RESV_ALIASES  = {'uber_reservation', 'uber_resv', 'uber予約'}
UBER_PROMO_ALIASES = {'uber_promo', 'uber_promotion', 'uberプロモーション'}

def is_uber_tip(pm_alias: str, cpm_alias: str, note: str, comment: str) -> bool:
    text = f"{note or ''} {comment or ''}"
    has_uber = ('uber' in (pm_alias or '')) or ('uber' in (cpm_alias or ''))
    if (pm_alias in UBER_TIP_ALIASES) or (cpm_alias in UBER_TIP_ALIASES):
        return True
    return has_uber and bool(TIP_PAT.search(text))

def is_uber_resv(pm_alias: str, cpm_alias: str, note: str, comment: str) -> bool:
    text = f"{note or ''} {comment or ''}"
    has_uber = ('uber' in (pm_alias or '')) or ('uber' in (cpm_alias or ''))
    if (pm_alias in UBER_RESV_ALIASES) or (cpm_alias in UBER_RESV_ALIASES):
        return True
    return has_uber and bool(RESV_PAT.search(text))

def is_uber_promo(pm_alias: str, cpm_alias: str, note: str, comment: str) -> bool:
    text = f"{note or ''} {comment or ''}"
    has_uber = ('uber' in (pm_alias or '')) or ('uber' in (cpm_alias or ''))
    if (pm_alias in UBER_PROMO_ALIASES) or (cpm_alias in UBER_PROMO_ALIASES):
        return True
    return has_uber and bool(PROMO_PAT.search(text))



# ========= 软预填（不落库，仅用于渲染初值） =========
def _safe_as_time(val):
    try:
        if val is None:
            return None
        if hasattr(val, "time") and callable(getattr(val, "time")):
            return val.time()
        if hasattr(val, "hour") and hasattr(val, "minute") and not hasattr(val, "date"):
            return val
        s = str(val).strip()
        if ":" in s:
            h, m = s.split(":", 1)
            h = int(h); m = int(m)
            if 0 <= h < 24 and 0 <= m < 60:
                from datetime import time as _t
                return _t(h, m)
    except Exception:
        pass
    return None

def _prefill_report_without_fk(report):
    """
    预填规则（只用实际值，不用计划值）：
    - 车辆：取当天该司机任一预约的 vehicle
    - 出勤：若为空，取当天所有预约中最早的 actual_departure
    - 退勤：若为空，取当天所有预约中最晚的 actual_return
      ▶ 若没有 actual_return，则保持空（绝不再用 end_time 回填）
    """
    try:
        user = getattr(getattr(report, "driver", None), "user", None)
        the_date = getattr(report, "date", None)
        if not user or not the_date:
            return

        # 当天所有覆盖该日期的预约（含跨天）
        qs = (Reservation.objects
              .filter(driver=user, date__lte=the_date, end_date__gte=the_date)
              .select_related("vehicle")
              .order_by("date", "start_time"))
        if not qs.exists():
            return

        # 车辆：缺就取第一条有车的
        if not getattr(report, "vehicle_id", None):
            for r in qs:
                v = getattr(r, "vehicle", None)
                if v:
                    report.vehicle = v
                    break

        # 出勤：仅取“实际出库”中最早的一个
        if getattr(report, "clock_in", None) in (None, ""):
            actual_deps = []
            for r in qs:
                ad = getattr(r, "actual_departure", None)
                if ad:
                    t = _safe_as_time(ad)
                    if t:
                        actual_deps.append(t)
            if actual_deps:
                report.clock_in = sorted(actual_deps)[0]

        # 退勤：仅取“实际入库”中最晚的一个；没有就保持空
        if getattr(report, "clock_out", None) in (None, ""):
            actual_returns = []
            for r in qs:
                ar = getattr(r, "actual_return", None)
                if ar:
                    t = _safe_as_time(ar)
                    if t:
                        actual_returns.append(t)
            if actual_returns:
                report.clock_out = sorted(actual_returns)[-1]
            # else: 不再用 end_time 填充，保持为空
    except Exception as e:
        debug_print("SOFT_PREFILL error:", e)

# ========= 小工具 =========
BASE_BREAK_MINUTES = 20
DEBUG_PRINT_ENABLED = True
if getattr(settings, "DEBUG", False):
    print("🔥 views.py 加载 OK")

def _to_int0(v):
    try:
        if v in ("", None):
            return 0
        return int(v)
    except (TypeError, ValueError):
        return 0

# 兼容旧代码里用到的 _to_int
_to_int = _to_int0

def _minutes_from_timedelta(td):
    if not td:
        return 0
    try:
        return int(td.total_seconds() // 60)
    except Exception:
        return 0

NIGHT_END_MIN = 5 * 60  # 05:00

def _sorted_items_qs(report):
    safe_ride = Coalesce(NullIf(F('ride_time'), Value('')), Value('00:00'))
    return (
        report.items
        .annotate(
            _safe_ride=safe_ride,
            _hour=Cast(Substr(F('_safe_ride'), 1, 2), IntegerField()),
            _minute=Cast(Substr(F('_safe_ride'), 4, 2), IntegerField()),
        )
        .annotate(_total_min=F('_hour') * 60 + F('_minute'))
        .annotate(
            _minutes_for_sort=ExpressionWrapper(
                F('_total_min') + Case(
                    When(_total_min__lt=NIGHT_END_MIN, then=Value(24 * 60)),
                    default=Value(0),
                ),
                output_field=IntegerField(),
            )
        )
        .order_by('_minutes_for_sort', 'id')
    )

def to_aware_dt(base_date, value, *, base_clock_in=None, tz=None):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, dtime):
        dt = datetime.combine(base_date, value)
    elif isinstance(value, str):
        s = value.strip()
        if not s or ":" not in s:
            return None
        try:
            h, m = map(int, s.split(":", 1))
        except Exception:
            return None
        dt = datetime.combine(base_date, dtime(hour=h, minute=m))
    else:
        return None

    if base_clock_in:
        ci = base_clock_in.time() if isinstance(base_clock_in, datetime) else base_clock_in
        if isinstance(ci, dtime) and dt.time() < ci:
            dt += timedelta(days=1)

    tz = tz or timezone.get_current_timezone()
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, tz)
    return dt

# === [SYNC UTILS START] 日报 <-> 预约 同步工具（在本文件内，不新建模块） ===
def _reservation_plan_window(reservation):
    """
    将 Reservation 的 (date, start_time) / (end_date, end_time)
    组合成本地时区 datetime 的计划窗口。
    """
    s = to_aware_dt(reservation.date, reservation.start_time)
    e = to_aware_dt(reservation.end_date, reservation.end_time, base_clock_in=reservation.start_time)
    return s, e


def _find_best_reservation_for_report(report, in_dt, out_dt):
    """
    在同一司机（report.driver.user）、同一车辆（若选择了车辆）、
    以 report.date 为中心 前后各 1 天 的范围内，选一条“最匹配”的预约：
      - 同时有 in/out：选“重叠时长最大”的预约
      - 只有一个时间点：选“距离最近”的预约
    """
    driver_user = getattr(getattr(report, "driver", None), "user", None)
    if not driver_user:
        return None

    qs = Reservation.objects.filter(
        driver=driver_user,
        date__lte=report.date + timedelta(days=1),
        end_date__gte=report.date - timedelta(days=1),
    )
    if getattr(report, "vehicle_id", None):
        qs = qs.filter(vehicle_id=report.vehicle_id)

    if not qs.exists():
        return None

    def overlap_or_gap(r):
        s, e = _reservation_plan_window(r)
        if in_dt and out_dt:
            a, b = max(s, in_dt), min(e, out_dt)
            overlap = max(timedelta(0), b - a)
            # 负数表示“更差”，用于排序（重叠越大越好）
            return (0, -overlap.total_seconds())
        else:
            t = in_dt or out_dt
            if s <= t <= e:
                gap = 0
            else:
                gap = min(abs(t - s), abs(t - e)).total_seconds()
            return (1, gap)

    # 按 (模式, 指标) 排序：模式 0(有重叠) 优于 1(只看距离)；指标越小越好
    best = sorted(qs, key=overlap_or_gap)[0]
    return best


def _sync_reservation_actual_for_report(report, old_clock_in, old_clock_out):
    """
    只在“从空到有”的场景下，同步 Reservation.actual_departure / actual_return。
    若预约里已有实际时间，则不覆盖。
    """
    # 判断是否“从空到有”
    filled_in_from_empty  = (not old_clock_in)  and bool(getattr(report, "clock_in",  None))
    filled_out_from_empty = (not old_clock_out) and bool(getattr(report, "clock_out", None))
    if not (filled_in_from_empty or filled_out_from_empty):
        return

    # 计算当天的 aware datetime（退勤相对出勤自动跨天）
    in_dt  = to_aware_dt(report.date, report.clock_in)  if getattr(report, "clock_in",  None) else None
    out_dt = to_aware_dt(report.date, report.clock_out, base_clock_in=in_dt) if getattr(report, "clock_out", None) else None

    reservation = _find_best_reservation_for_report(report, in_dt, out_dt)
    if not reservation:
        return

    updated_fields = []
    if filled_in_from_empty and in_dt and getattr(reservation, "actual_departure", None) in (None, ""):
        reservation.actual_departure = in_dt
        updated_fields.append("actual_departure")

    if filled_out_from_empty and out_dt and getattr(reservation, "actual_return", None) in (None, ""):
        reservation.actual_return = out_dt
        updated_fields.append("actual_return")

    if updated_fields:
        reservation.save(update_fields=updated_fields)
# === [SYNC UTILS END] ===

def check_module_permission(user, perm_key: str) -> bool:
    try:
        if not getattr(user, "is_authenticated", False):
            return False
        if getattr(user, "is_superuser", False):
            return True

        APP_LABEL = "dailyreport"
        key = (perm_key or "").strip().lower()

        candidates = [
            f"{APP_LABEL}.{key}",
            f"{APP_LABEL}.can_{key}",
            f"{APP_LABEL}.is_{key}",
            key,
        ]
        for perm in candidates:
            try:
                if user.has_perm(perm):
                    return True
            except Exception:
                pass

        try:
            if user.has_module_perms(APP_LABEL):
                return True
        except Exception:
            pass

        try:
            group_names = {g.name.strip().lower() for g in user.groups.all()}
            if key in group_names or f"{APP_LABEL}:{key}" in group_names:
                return True
        except Exception:
            pass

        return False
    except Exception:
        return False


# ==== BEGIN REPLACE: is_dailyreport_admin (dailyreport/views.py) ====
def is_dailyreport_admin(user):
    """
    日报系统管理权限：
    - 超级用户
    - UserProfile.is_dispatch_admin = True  （配车系统管理员）
    - UserProfile.is_dailyreport_admin = True（日报管理系统管理员）
    （员工台账系统管理员 is_staffbook_admin 不再自动拥有日报权限）

    同时保留基于 permission 的判断（dailyreport_admin / dailyreport），方便以后扩展。
    """
    try:
        # 未登录一律不允许
        if not getattr(user, "is_authenticated", False):
            return False

        # 超级用户永远允许
        if getattr(user, "is_superuser", False):
            return True

        # 先看 UserProfile 上的布尔位
        profile = getattr(user, "userprofile", None)
        if profile is not None:
            if getattr(profile, "is_dispatch_admin", False) or getattr(
                profile, "is_dailyreport_admin", False
            ):
                return True

        # 再看基于权限字符串的判断（沿用原来的灵活机制）
        if (
            check_module_permission(user, "dailyreport_admin")
            or check_module_permission(user, "dailyreport")
        ):
            return True

        # 普通 is_staff 不再赋予日报管理权
        return False
    except Exception:
        # 出错时保守处理：只给超管
        return bool(getattr(user, "is_superuser", False))
# ==== END REPLACE: is_dailyreport_admin (dailyreport/views.py) ====


dailyreport_admin_required = user_passes_test(is_dailyreport_admin)



def get_active_drivers(month_obj=None, keyword=None):
    qs = Driver.objects.all()
    if month_obj is None:
        month_obj = date.today()

    year = month_obj.year
    month = month_obj.month
    from datetime import date as _date
    from calendar import monthrange as _monthrange
    first_day = _date(year, month, 1)
    last_day = _date(year, month, _monthrange(year, month)[1])

    try:
        qs = qs.filter(
            Q(hire_date__lte=last_day)
            & (Q(resigned_date__isnull=True) | Q(resigned_date__gte=first_day))
        )
    except Exception:
        pass

    if hasattr(Driver, 'is_active'):
        try:
            qs = qs.filter(is_active=True)
        except Exception:
            pass

    if keyword:
        try:
            qs = qs.filter(Q(name__icontains=keyword) | Q(code__icontains=keyword))
        except Exception:
            pass

    return qs.order_by('name')

# ========= 基础视图 =========
@user_passes_test(is_dailyreport_admin)
def dailyreport_create(request):
    if request.method == 'POST':
        form = DriverDailyReportForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dailyreport:dailyreport_list')
    else:
        form = DriverDailyReportForm()
    return render(request, 'dailyreport/driver_dailyreport_edit.html', {'form': form})


PREFIX = "items"   # ✅ 前后端统一的前缀

ReportItemFormSet = inlineformset_factory(
    DriverDailyReport,
    DriverDailyReportItem,
    form=DriverDailyReportItemForm,
    extra=0,
    can_delete=True,     # ✅ 允许删除
    max_num=40,
)


def dailyreport_edit(request, pk):
    report = get_object_or_404(DriverDailyReport, pk=pk)

    if request.method == 'POST':
        form = DriverDailyReportForm(request.POST, instance=report)
        formset = ReportItemFormSet(request.POST, instance=report, prefix=PREFIX)

        if form.is_valid() and formset.is_valid():
            inst = form.save(commit=False)
            inst.edited_by = request.user

                        # ===== [PATCH PAYROLL SAVE-GUARD BEGIN] =====
            # payroll_* は JS が hidden に書き込むが、
            # POST欠落/空送信でも None を入れない & 既存値を守る
            PAYROLL_FIELDS = [
                "payroll_total",
                "payroll_bd_sales",
                "payroll_bd_advance",
                "payroll_bd_etc_refund",
                "payroll_bd_over_short_to_driver",
                "payroll_bd_over_short_to_company",
            ]

            for f in PAYROLL_FIELDS:
                # POSTに含まれていなければ「既存値を保持」
                if f not in request.POST:
                    setattr(inst, f, getattr(report, f, 0) or 0)
                    continue

                # POSTにあるが form が None を作った場合は既存値/0に寄せる
                if getattr(inst, f, None) is None:
                    setattr(inst, f, getattr(report, f, 0) or 0)
            # ===== [PATCH PAYROLL SAVE-GUARD END] =====


            inst.save()

            # 关键：一句话就够了（增/改/删 都在这里完成）
            formset.instance = inst
            formset.save()   # ✅ 会自动删除勾选 DELETE 的旧行

            # ===== [PATCH PAYROLL SAVE BEGIN] 給与計算用（表示）を落庫 =====
            def _to_int0(v):
                try:
                    return int(v)
                except (TypeError, ValueError):
                    return 0

            # JSが hidden に書き込んだ値を保存する（無ければ 0）
            inst.payroll_total = _to_int0(request.POST.get("payroll_total"))

            inst.payroll_bd_sales = _to_int0(request.POST.get("payroll_bd_sales"))
            inst.payroll_bd_advance = _to_int0(request.POST.get("payroll_bd_advance"))
            inst.payroll_bd_etc_refund = _to_int0(request.POST.get("payroll_bd_etc_refund"))
            inst.payroll_bd_over_short_to_driver = _to_int0(request.POST.get("payroll_bd_over_short_to_driver"))
            inst.payroll_bd_over_short_to_company = _to_int0(request.POST.get("payroll_bd_over_short_to_company"))

            inst.save(update_fields=[
                "payroll_total",
                "payroll_bd_sales",
                "payroll_bd_advance",
                "payroll_bd_etc_refund",
                "payroll_bd_over_short_to_driver",
                "payroll_bd_over_short_to_company",
            ])
            # ===== [PATCH PAYROLL SAVE END] =====


            messages.success(request, "保存成功！")
            return redirect('dailyreport:dailyreport_edit', pk=inst.pk)
        else:
            # 🔥 调试：把错误打到控制台 或 log
            print("【DEBUG】日报主表错误：", form.errors)
            print("【DEBUG】明细行错误：", formset.errors)
            messages.error(request, "保存失败，请检查输入内容")
    else:
        form = DriverDailyReportForm(instance=report)
        formset = ReportItemFormSet(instance=report, prefix=PREFIX)  # ✅ GET 同样用 prefix

    # 模板需要的其它上下文按你现有的来，这里只保证能渲染
    return render(request, 'dailyreport/driver_dailyreport_edit.html', {
        'form': form,
        'formset': formset,
        'report': report,
        'driver': getattr(report, 'driver', None),
        'is_edit': True,
    })

    
from django.views.decorators.http import require_POST, require_http_methods
# 如果上面没引入 user_passes_test / 模型，也一并确认
from .models import DriverDailyReportItem, Driver
@user_passes_test(is_dailyreport_admin)
@require_POST
def dailyreport_item_delete(request, item_id):
    item = get_object_or_404(DriverDailyReportItem, pk=item_id)
    report_id = item.report_id
    item.delete()
    messages.success(request, "已删除 1 条明细。")
    return redirect('dailyreport:dailyreport_edit', pk=report_id)

@login_required
def sales_thanks(request):
    return render(request, 'dailyreport/sales_thanks.html')

@user_passes_test(is_dailyreport_admin)
def dailyreport_delete_for_driver(request, driver_id, pk):
    driver = get_object_or_404(Driver, pk=driver_id)
    report = get_object_or_404(DriverDailyReport, pk=pk, driver=driver)
    if request.method == "POST":
        report.delete()
        messages.success(request, "已删除该日报记录。")
        return redirect('dailyreport:driver_dailyreport_month', driver_id=driver.id)
    return render(request, 'dailyreport/dailyreport_confirm_delete.html', {
        'report': report,
        'driver': driver,
    })

@login_required
def dailyreport_list(request):
    if request.user.is_staff:
        reports = DriverDailyReport.objects.all().order_by('-date')
    else:
        reports = DriverDailyReport.objects.filter(driver=request.user).order_by('-date')
    return render(request, 'dailyreport/dailyreport_list.html', {'reports': reports})


def _filter_by_driver_id(qs, request):
    """
    若 GET 里带了 ?driver_id=XX，则按司机过滤 QuerySet。
    兼容空/非法输入（直接忽略）。
    """
    driver_id = (request.GET.get("driver_id") or "").strip()
    if not driver_id:
        return qs, None
    try:
        did = int(driver_id)
    except (TypeError, ValueError):
        return qs, None

    qs = qs.filter(driver_id=did)

    # 为了在文件名里显示司机名，尽量取一个 Driver 对象（失败就返回 None）
    try:
        d = Driver.objects.only("id", "name", "driver_code").get(id=did)
    except Exception:
        d = None
    return qs, d

@user_passes_test(is_dailyreport_admin)
def export_dailyreports_excel(request, year, month):
    """
    FINAL: Excel 导出（3类Sheet固定存在）
      Sheet① 索引
      Sheet② 每日明细（每天一个sheet）
      Sheet③ 集計（区间/月份按司机汇总 + 总出勤时长）
    """
    from collections import defaultdict
    from datetime import datetime, timedelta, date, time as dtime
    from decimal import Decimal, ROUND_HALF_UP
    from io import BytesIO
    from urllib.parse import quote

    from django.db.models import Case, When, F, DateField, ExpressionWrapper
    from django.http import FileResponse, HttpResponse

    try:
        import xlsxwriter
    except ModuleNotFoundError:
        return HttpResponse(
            "XlsxWriter 未安装。请在虚拟环境中运行：pip install XlsxWriter",
            status=500
        )

    # =========================================================
    # 参数
    # =========================================================
    FEE_RATE = Decimal("0.05")

    def fee_calc(x: int) -> int:
        if not x:
            return 0
        return int((Decimal(x) * FEE_RATE).quantize(Decimal("1"), rounding=ROUND_HALF_UP))

    # ながし現金：非貸切 + 现金类 payment_method
    NAGASHI_CASH_METHODS = {"cash", "uber_cash", "didi_cash", "go_cash"}

    # =========================================================
    # 区间导出支持
    # =========================================================
    q_from = (request.GET.get("from") or "").strip()
    q_to   = (request.GET.get("to") or "").strip()
    date_range = None
    if q_from and q_to:
        try:
            date_from = datetime.strptime(q_from, "%Y-%m-%d").date()
            date_to   = datetime.strptime(q_to, "%Y-%m-%d").date()
            if date_from > date_to:
                return HttpResponse("開始日必须早于/等于終了日", status=400)
            date_range = (date_from, date_to)
        except ValueError:
            return HttpResponse("日期格式应为 YYYY-MM-DD", status=400)

    # =========================================================
    # 勤务日规则（06:00 前算前一天）
    # =========================================================
    work_date_expr = Case(
        When(
            clock_in__lt=dtime(6, 0),
            then=ExpressionWrapper(F("date") - timedelta(days=1), output_field=DateField()),
        ),
        default=F("date"),
        output_field=DateField(),
    )

    qs = (
        DriverDailyReport.objects
        .annotate(work_date=work_date_expr)
        .select_related("driver")
        .prefetch_related("items")
    )

    if date_range:
        reports = qs.filter(date__range=date_range).order_by("work_date", "driver__name")
        range_from, range_to = date_range
    else:
        reports = qs.filter(work_date__year=year, work_date__month=month).order_by("work_date", "driver__name")
        range_from = date(year, month, 1)
        range_to = (range_from.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    # =========================================================
    # 按日期分组（Sheet②）
    # =========================================================
    by_date = defaultdict(list)
    for r in reports:
        by_date[getattr(r, "work_date", None) or r.date].append(r)

    # =========================================================
    # Excel 初始化
    # =========================================================
    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {"in_memory": True})

    fmt_header = wb.add_format({"bold": True, "align": "center", "valign": "vcenter", "border": 1, "bg_color": "#DDDDDD"})
    fmt_border = wb.add_format({"border": 1})
    fmt_yen = wb.add_format({"border": 1, "align": "right", "num_format": "¥#,##0"})
    fmt_hour = wb.add_format({"border": 1, "align": "right", "num_format": "#,##0.00"})
    fmt_note = wb.add_format({"italic": True, "font_color": "#555555"})
    fmt_ng = wb.add_format({"border": 1, "bg_color": "#FFC7CE", "font_color": "#9C0006"})

    # =========================================================
    # Sheet① 索引
    # =========================================================
    ws_index = wb.add_worksheet("索引")
    ws_index.write_row(0, 0, ["日期", "件数"], fmt_header)
    idx_r = 1
    for d, reps in sorted(by_date.items()):
        ws_index.write(idx_r, 0, d.strftime("%Y-%m-%d"), fmt_border)
        ws_index.write_number(idx_r, 1, len(reps), fmt_border)
        idx_r += 1

    # =========================================================
    # Sheet② 每日明细（事实账）— 每天一个sheet（不可删除）
    # =========================================================
    headers_daily = [
        "社員番号","従業員","出勤","退勤",
        "ながし現金","貸切現金","ETC",
        "入金予定","実入金額","給油","CHECK","NG原因",
        "貸切未収",
        "Uber","手数料",
        "PayPay","手数料",
        "DiDi","手数料",
        "水揚合計","税抜","消費税","過不足",
    ]

    def compute_daily_row(r):
        nagashi_cash = 0
        charter_cash = 0
        charter_uncol = 0

        amt = {"uber": 0, "paypay": 0, "didi": 0}

        for it in r.items.all():
            meter = int(getattr(it, "meter_fee", 0) or 0)
            is_charter = bool(getattr(it, "is_charter", False))

            if not is_charter:
                if getattr(it, "payment_method", "") in NAGASHI_CASH_METHODS:
                    nagashi_cash += meter
            else:
                charter_amt = int(getattr(it, "charter_amount_jpy", 0) or 0)
                if getattr(it, "charter_payment_method", "") == "cash":
                    charter_cash += charter_amt
                else:
                    charter_uncol += charter_amt

            pm = getattr(it, "payment_method", "")
            if pm in amt:
                amt[pm] += meter

        uber_fee   = fee_calc(amt["uber"])
        paypay_fee = fee_calc(amt["paypay"])
        didi_fee   = fee_calc(amt["didi"])

        expected = nagashi_cash + charter_cash + int(getattr(r, "etc_collected", 0) or 0)
        deposit  = int(getattr(r, "deposit_amount", 0) or 0)
        diff     = int(getattr(r, "deposit_difference", 0) or 0)

        calc_delta = deposit - expected   # ✅ 就是这一行
        check    = "OK" if (deposit - expected) == diff else "NG"
        # ===== NG原因 =====
        if check == "OK":
            ng_reason = ""
        elif calc_delta < 0:
            ng_reason = f"入金不足（{calc_delta}）"
        else:
            ng_reason = f"入金过多（+{calc_delta}）"

        water = nagashi_cash + charter_cash + charter_uncol
        tax_ex = int((Decimal(water) / Decimal("1.1")).quantize(Decimal("1"), rounding=ROUND_HALF_UP)) if water else 0
        tax = water - tax_ex

        return [
            getattr(getattr(r, "driver", None), "driver_code", "") or "",
            getattr(getattr(r, "driver", None), "name", "") or "",
            r.clock_in.strftime("%H:%M") if getattr(r, "clock_in", None) else "",
            r.clock_out.strftime("%H:%M") if getattr(r, "clock_out", None) else "",
            nagashi_cash,
            charter_cash,
            int(getattr(r, "etc_collected", 0) or 0),
            expected,
            deposit,
            int(getattr(r, "fuel_amount", 0) or 0),
            check,
            ng_reason,      # ✅ 现在会写内容
            charter_uncol,
            amt["uber"], uber_fee,
            amt["paypay"], paypay_fee,
            amt["didi"], didi_fee,
            water,
            tax_ex,
            tax,
            diff,
        ]

    for d, reps in sorted(by_date.items()):
        ws = wb.add_worksheet(d.strftime("%Y-%m-%d"))
        ws.merge_range(
            0, 0, 0, len(headers_daily) - 1,
            "※ 入金予定 = ながし現金 + 貸切現金 + 実際ETC（給油は含めない）",
            fmt_note
        )
        ws.write_row(1, 0, headers_daily, fmt_header)
        ws.freeze_panes(2, 4)

        row = 2
        for rep in reps:
            vals = compute_daily_row(rep)
            for c, v in enumerate(vals):
                if isinstance(v, int):
                    ws.write_number(row, c, v, fmt_yen)
                else:
                    ws.write(row, c, v, fmt_border)
            row += 1

        check_col = headers_daily.index("CHECK")
        if row - 1 >= 2:
            ws.conditional_format(
                2, check_col, row - 1, check_col,
                {"type": "text", "criteria": "containing", "value": "NG", "format": fmt_ng}
            )

    # =========================================================
    # Sheet③ 集計（按司机 / 只 SUM + 总出勤时长）
    # =========================================================
    summary = defaultdict(lambda: {
        "days": set(),
        "work_minutes": 0,
        "nagashi": 0,
        "charter_cash": 0,
        "charter_uncol": 0,
        "etc": 0,
        "fuel": 0,
        "deposit": 0,
        "diff_pos": 0,
        "diff_neg": 0,
        "uber": 0, "uber_fee": 0,
        "paypay": 0, "paypay_fee": 0,
        "didi": 0, "didi_fee": 0,
    })

    for r in reports:
        s = summary[r.driver_id]
        base_date = getattr(r, "work_date", None) or r.date
        s["days"].add(base_date)

        # 出勤分钟（time不能直接相减，必须先拼 datetime）
        if r.clock_in and r.clock_out:
            dt_in = datetime.combine(base_date, r.clock_in)
            dt_out = datetime.combine(base_date, r.clock_out)
            if dt_out < dt_in:
                dt_out += timedelta(days=1)  # 跨日
            minutes = int((dt_out - dt_in).total_seconds() // 60)
            s["work_minutes"] += max(minutes, 0)

        s["etc"] += int(getattr(r, "etc_collected", 0) or 0)
        s["fuel"] += int(getattr(r, "fuel_amount", 0) or 0)
        s["deposit"] += int(getattr(r, "deposit_amount", 0) or 0)

        diff = int(getattr(r, "deposit_difference", 0) or 0)
        if diff > 0:
            s["diff_pos"] += diff
        elif diff < 0:
            s["diff_neg"] += abs(diff)

        for it in r.items.all():
            meter = int(getattr(it, "meter_fee", 0) or 0)
            is_charter = bool(getattr(it, "is_charter", False))

            if not is_charter:
                if getattr(it, "payment_method", "") in NAGASHI_CASH_METHODS:
                    s["nagashi"] += meter
            else:
                charter_amt = int(getattr(it, "charter_amount_jpy", 0) or 0)
                if getattr(it, "charter_payment_method", "") == "cash":
                    s["charter_cash"] += charter_amt
                else:
                    s["charter_uncol"] += charter_amt

            pm = getattr(it, "payment_method", "")
            if pm == "uber":
                s["uber"] += meter
                s["uber_fee"] += fee_calc(meter)
            elif pm == "paypay":
                s["paypay"] += meter
                s["paypay_fee"] += fee_calc(meter)
            elif pm == "didi":
                s["didi"] += meter
                s["didi_fee"] += fee_calc(meter)

    ws_sum = wb.add_worksheet(f"{range_from}~{range_to}(集計)")
    sum_headers = [
        "社員番号","従業員","日数","総出勤時間(H)",
        "ながし","貸切現金","貸切未収",
        "ETC","給油","入金額",
        "過不足＋","過不足−",
        "Uber","手数料",
        "PayPay","手数料",
        "DiDi","手数料",
    ]
    ws_sum.write_row(0, 0, sum_headers, fmt_header)
    ws_sum.freeze_panes(1, 2)

    col_days = sum_headers.index("日数")
    col_hours = sum_headers.index("総出勤時間(H)")

    row = 1
    for driver_id, s in summary.items():
        # 注意：vals 必须无条件定义，且只在本循环内使用（防 UnboundLocalError）
        drv = Driver.objects.get(id=driver_id)

        vals = [
            getattr(drv, "driver_code", "") or "",
            getattr(drv, "name", "") or "",
            len(s["days"]),
            round(s["work_minutes"] / 60, 2),
            s["nagashi"],
            s["charter_cash"],
            s["charter_uncol"],
            s["etc"],
            s["fuel"],
            s["deposit"],
            s["diff_pos"],
            s["diff_neg"],
            s["uber"], s["uber_fee"],
            s["paypay"], s["paypay_fee"],
            s["didi"], s["didi_fee"],
        ]

        for c, v in enumerate(vals):
            if c == col_days:
                ws_sum.write_number(row, c, int(v), fmt_border)   # 日数：非金额
            elif c == col_hours:
                ws_sum.write_number(row, c, float(v), fmt_hour)   # 工时：非金额
            elif isinstance(v, int):
                ws_sum.write_number(row, c, int(v), fmt_yen)      # 金额
            else:
                ws_sum.write(row, c, v, fmt_border)

        row += 1

    wb.close()
    output.seek(0)

    filename = f"{range_from}~{range_to}_全員毎日集計.xlsx"
    return FileResponse(
        output,
        as_attachment=True,
        filename=quote(filename),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )




# ========= 合计辅助 =========
def _normalize(val: str) -> str:
    if not val:
        return ''
    v = str(val).strip().lower()
    mapping = {
        'jpy_cash': 'jpy_cash', 'rmb_cash': 'rmb_cash',
        'self_wechat': 'self_wechat', 'boss_wechat': 'boss_wechat',
        'to_company': 'to_company', 'bank_transfer': 'bank_transfer',
        '--------': '', '------': '', '': '',
        '現金': 'jpy_cash', '现金': 'jpy_cash', '日元現金': 'jpy_cash', '日元现金': 'jpy_cash',
        '人民幣現金': 'rmb_cash', '人民币现金': 'rmb_cash',
        '自有微信': 'self_wechat', '老板微信': 'boss_wechat',
        '公司回收': 'to_company', '会社回収': 'to_company', '公司结算': 'to_company',
        '銀行振込': 'bank_transfer', 'bank': 'bank_transfer',
    }
    return mapping.get(v, v)

def _totals_of(items):
    """编辑页用的那套：メータのみ + 貸切现金/未収/不明 + sales_total"""
    meter_only = Decimal('0')
    charter_cash = Decimal('0')
    charter_uncol = Decimal('0')
    charter_unknown = Decimal('0')

    for it in items:
        if getattr(it, 'is_charter', False):
            amt = Decimal(getattr(it, 'charter_amount_jpy', 0) or 0)
            if amt <= 0:
                continue
            method = _normalize(getattr(it, 'charter_payment_method', ''))
            if method in {'jpy_cash', 'rmb_cash', 'self_wechat', 'boss_wechat'}:
                charter_cash += amt
            elif method in {'to_company', 'bank_transfer', ''}:
                charter_uncol += amt
            else:
                charter_unknown += amt
        else:
            # 只有有 payment_method 的才算メータのみ
            if getattr(it, 'payment_method', None):
                meter_only += Decimal(it.meter_fee or 0)

    sales_total = meter_only + charter_cash + charter_uncol + charter_unknown
    return {
        'meter_only_total': meter_only,
        'charter_cash_total': charter_cash,
        'charter_uncollected_total': charter_uncol,
        'charter_unknown_total': charter_unknown,
        'sales_total': sales_total,
    }

# ========= 月视图 =========
@user_passes_test(is_dailyreport_admin)
def driver_dailyreport_month(request, driver_id):
    from decimal import Decimal  # 以防上面没导入

    driver = get_object_or_404(Driver, id=driver_id)

    # 対象月
    month_str = request.GET.get("month", "")
    try:
        month = datetime.strptime(month_str, "%Y-%m").date().replace(day=1)
        month_str = month.strftime("%Y-%m")
    except Exception:
        month = timezone.localdate().replace(day=1)
        month_str = month.strftime("%Y-%m")

    reports_qs = (
        DriverDailyReport.objects
        .filter(driver=driver, date__year=month.year, date__month=month.month)
        .order_by("-date")
        .prefetch_related("items")
    )

    report_list = []

    def _amount_for_item(it):
        """
        与 Excel / JS 一致：
          - 貸切行用 charter_amount_jpy
          - それ以外は meter_fee
        """
        if getattr(it, "is_charter", False):
            return Decimal(getattr(it, "charter_amount_jpy", 0) or 0)
        return Decimal(getattr(it, "meter_fee", 0) or 0)

    for r in reports_qs:
        items = list(r.items.all())

        # ① 用 _totals_of 算 メータのみ / 貸切現金 / 貸切未収 / 未分類
        base_totals = _totals_of(items)

        # ② Uber 予約 / チップ / プロモ
        uber_resv = uber_tip = uber_promo = Decimal("0")
        for it in items:
            pm_alias = (getattr(it, "payment_method", "") or "").strip().lower()
            cpm_alias = (getattr(it, "charter_payment_method", "") or "").strip().lower()
            note = getattr(it, "note", "") or ""
            comment = getattr(it, "comment", "") or ""

            if is_uber_resv(pm_alias, cpm_alias, note, comment):
                uber_resv += _amount_for_item(it)
            elif is_uber_tip(pm_alias, cpm_alias, note, comment):
                uber_tip += _amount_for_item(it)
            elif is_uber_promo(pm_alias, cpm_alias, note, comment):
                uber_promo += _amount_for_item(it)

        uber_total = uber_resv + uber_tip + uber_promo

        # ③ 売上合計（按 A 案）
        # === BEGIN PATCH: 月一览の売上合計を編集ページと同じ口径に揃える ===
        # _totals_of() からの値（ここには Uber予約/チップ/プロモ も含まれている）
        meter_only_raw   = base_totals["meter_only_total"]
        charter_cash     = base_totals["charter_cash_total"]
        charter_uncol    = base_totals["charter_uncollected_total"]
        charter_unknown  = base_totals["charter_unknown_total"]
        base_sales_total = base_totals["sales_total"]  # = meter_only_raw + 貸切現金 + 貸切未収 + 未分類

        # 編集ページの表示に合わせて：
        # 「メータのみ」は Uber予約/チップ/プロモ を差し引いた値にする
        meter_only_without_uber = meter_only_raw - uber_total
        if meter_only_without_uber < 0:
            meter_only_without_uber = Decimal("0")

        # 月一览に渡す値をセット
        r.meter_only_total          = meter_only_without_uber
        r.charter_cash_total        = charter_cash
        r.charter_uncollected_total = charter_uncol
        r.charter_unknown_total     = charter_unknown

        r.uber_reservation_total = uber_resv
        r.uber_tip_total         = uber_tip
        r.uber_promotion_total   = uber_promo

        # 売上合計は _totals_of() の sales_total をそのまま使う
        # （ここには既に Uber予約/チップ/プロモ が含まれているので、二重に足さない）
        r.total_all = base_sales_total
        # === END PATCH ===

        report_list.append(r)

    context = {
        "driver": driver,
        "reports": report_list,
        "selected_month": month_str,
        "selected_date": request.GET.get("date", ""),
        "today": timezone.localdate(),
        "month": month,
    }
    return render(request, "dailyreport/driver_dailyreport_month.html", context)


# ========= 选择器 & 直接创建 =========
@user_passes_test(is_dailyreport_admin)
def dailyreport_add_selector(request, driver_id):
    from datetime import date as _date
    driver = get_object_or_404(Driver, pk=driver_id)

    month_str = request.GET.get("month")
    try:
        if month_str:
            target_year, target_month = map(int, month_str.split("-"))
            display_date = _date(target_year, target_month, 1)
        else:
            display_date = _date.today()
    except ValueError:
        display_date = _date.today()

    current_month = display_date.strftime("%Y-%m")
    num_days = monthrange(display_date.year, display_date.month)[1]
    all_dates = [_date(display_date.year, display_date.month, d) for d in range(1, num_days + 1)]

    reserved_dates = set()
    if driver.user:
        reserved_dates = set(
            Reservation.objects
            .filter(driver=driver.user, date__year=display_date.year, date__month=display_date.month)
            .values_list("date", flat=True)
        )

    calendar_dates = [{"date": d, "enabled": d in reserved_dates} for d in all_dates]

    if request.method == "POST":
        selected_date_str = request.POST.get("selected_date")
        try:
            selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "无效的日期")
            return redirect(request.path)

        if not driver.user or not Reservation.objects.filter(driver=driver.user, date=selected_date).exists():
            messages.warning(request, f"{selected_date.strftime('%Y年%m月%d日')} は出勤予約がありません。日報を作成できません。")
            return redirect(request.path + f"?month={current_month}")

        report, created = DriverDailyReport.objects.get_or_create(
            driver=driver,
            date=selected_date,
            defaults={"status": "pending"}
        )

        if created:
            res = (
                Reservation.objects
                .filter(driver=driver.user, date=selected_date)
                .order_by('start_time')
                .first()
            )
            if res:
                if res.vehicle:
                    report.vehicle = res.vehicle
                if res.actual_departure:
                    report.clock_in = timezone.localtime(res.actual_departure).time()
                if res.actual_return:
                    report.clock_out = timezone.localtime(res.actual_return).time()
                report.save()

        return redirect("dailyreport:driver_dailyreport_edit", driver_id=driver.id, report_id=report.id)

    return render(request, "dailyreport/driver_dailyreport_add.html", {
        "driver": driver,
        "current_month": display_date.strftime("%Y年%m月"),
        "year": display_date.year,
        "month": display_date.month,
        "calendar_dates": calendar_dates,
    })

@user_passes_test(is_dailyreport_admin)
def dailyreport_create_for_driver(request, driver_id):
    driver = get_driver_info(driver_id)
    if not driver:
        return render(request, 'dailyreport/not_found.html', status=404)

    if request.method == 'GET' and request.GET.get('date'):
        try:
            the_date = datetime.strptime(request.GET.get('date'), "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "无效的日期格式")
            return redirect('dailyreport:driver_dailyreport_month', driver_id=driver.id)

        existing = DriverDailyReport.objects.filter(driver=driver, date=the_date).first()
        if existing:
            return redirect('dailyreport:driver_dailyreport_edit', driver_id=driver.id, report_id=existing.id)

        new_report = DriverDailyReport.objects.create(driver=driver, date=the_date)
        return redirect('dailyreport:driver_dailyreport_edit', driver_id=driver.id, report_id=new_report.id)

    ReportItemFS = inlineformset_factory(
        DriverDailyReport, DriverDailyReportItem,
        form=DriverDailyReportItemForm,
        formset=RequiredReportItemFormSet,
        extra=0, can_delete=True, max_num=40,
    )

    if request.method == 'POST':
        report_form = DriverDailyReportForm(request.POST)
        formset = ReportItemFS(request.POST)

        if report_form.is_valid() and formset.is_valid():
            dailyreport = report_form.save(commit=False)
            dailyreport.driver = driver
            try:
                dailyreport.calculate_work_times()
            except Exception:
                pass

            cash_total = sum(
                item.cleaned_data.get('meter_fee') or 0
                for item in formset.forms
                if item.cleaned_data.get('payment_method') == 'cash'
                and not item.cleaned_data.get('DELETE', False)
            )
            deposit = dailyreport.deposit_amount or 0
            dailyreport.deposit_difference = deposit - cash_total

            dailyreport.save()
            formset.instance = dailyreport
            formset.save()

            messages.success(request, '新增日报成功')
            return redirect('dailyreport:driver_dailyreport_month', driver_id=driver.id)
        else:
            print("日报主表错误：", report_form.errors)
            print("明细表错误：", formset.errors)
    else:
        report_form = DriverDailyReportForm()
        formset = ReportItemFS()

    if request.method == 'POST' and formset.is_valid():
        data_iter = [f.cleaned_data for f in formset.forms if f.cleaned_data]
        totals = calculate_totals_from_formset(data_iter)
    else:
        data_iter = [f.instance for f in formset.forms]
        totals = calculate_totals_from_instances(data_iter)

    summary_keys = [
        ('meter', 'メーター(水揚)'),
        ('cash', '現金(ながし)'),
        ('uber', 'Uber'),
        ('didi', 'Didi'),
        ('credit', 'クレジ'),
        ('kyokushin', '京交信'),
        ('omron', 'オムロン(愛のタクシーチケット)'),
        ('kyotoshi', '京都市他'),
        ('qr', '扫码'),
    ]

    return render(request, 'dailyreport/driver_dailyreport_edit.html', {
        'form': report_form,
        'formset': formset,
        'driver': driver,
        'report': None,
        'is_edit': False,
        'summary_keys': summary_keys,
        'totals': totals,
    })

# ========= 编辑（员工） =========
@user_passes_test(is_dailyreport_admin)
def dailyreport_edit_for_driver(request, driver_id, report_id):
    driver = get_driver_info(driver_id)
    if not driver:
        return render(request, "dailyreport/not_found.html", status=404)

    # 防变量遮蔽：避免有人在函数内部把 DriverDailyReport 当作变量名赋值
    # 用 apps.get_model 以“字符串”方式获取模型，绕开名字遮蔽。
    from django.apps import apps
    DR = apps.get_model('dailyreport', 'DriverDailyReport')
    report = get_object_or_404(DR, pk=report_id, driver_id=driver_id)

    ReportItemFormSet = inlineformset_factory(
        DR,
        DriverDailyReportItem,
        form=DriverDailyReportItemForm,
        formset=RequiredReportItemFormSet,
        extra=0,
        can_delete=True,
        max_num=40,
    )

    if request.method == 'POST':
        post = request.POST.copy()

        if not post.get("vehicle") and report.vehicle_id:
            post["vehicle"] = str(report.vehicle_id)

        PM_ALIASES = {
            'company card': 'credit', 'Company Card': 'credit', '会社カード': 'credit',
            'company_card': 'credit', 'credit card': 'credit',
            'バーコード': 'qr', 'barcode': 'qr', 'bar_code': 'qr', 'qr_code': 'qr', 'qr': 'qr',
            '現金': 'cash', '现金': 'cash', 'cash(現金)': 'cash',
            'uber現金': 'uber_cash', 'didi現金': 'didi_cash', 'go現金': 'go_cash',
        }
        for k, v in list(post.items()):
            if k.endswith('-payment_method'):
                post[k] = PM_ALIASES.get(v, v)

        # ✅ 只交给表单“HH:MM”，表单验证通过后我们再拼成当天的 datetime 存库
        post['clock_in']  = _norm_hhmm(post.get('clock_in'))
        post['clock_out'] = _norm_hhmm(post.get('clock_out'))

        form = DriverDailyReportForm(post, instance=report)

        # 🔧 修复点：与模板/JS 一致的前缀，避免把旧行当“新增”
        formset = ReportItemFormSet(post, instance=report, prefix=PREFIX)

        if form.is_valid() and formset.is_valid():
            # >>> BEGIN DEBUG_ETC_FORMSET
            # 打印每行 ETC 相关字段，确认 POST 进来的值
            print("===== DEBUG ETC formset cleaned_data =====")
            for idx, f in enumerate(formset.forms):
                if not hasattr(f, "cleaned_data"):
                    continue
                cd = f.cleaned_data
                # 跳过被标记删除的行
                if cd.get("DELETE"):
                    continue
                print(
                    f"[ROW {idx}] "
                    f"id={cd.get('id')!r} "
                    f"etc_riding={cd.get('etc_riding')!r} "
                    f"etc_riding_charge_type={cd.get('etc_riding_charge_type')!r} "
                    f"etc_empty={cd.get('etc_empty')!r} "
                    f"etc_empty_charge_type={cd.get('etc_empty_charge_type')!r} "
                    f"etc_charge_type={cd.get('etc_charge_type')!r}"
                )
            print("===== END DEBUG ETC formset cleaned_data =====")
            # >>> END DEBUG_ETC_FORMSET

            
            # === 记录保存前的旧值 ===
            _old_in  = getattr(report, "clock_in",  None)
            _old_out = getattr(report, "clock_out", None)

            inst = form.save(commit=False)

            if not inst.vehicle_id and report.vehicle_id:
                inst.vehicle_id = report.vehicle_id

            break_input = (post.get("break_time_input") or "").strip()
            user_minutes = 0
            try:
                if ":" in break_input:
                    h, m = map(int, break_input.split(":", 1))
                    user_minutes = max(0, h * 60 + m)
                elif break_input:
                    user_minutes = max(0, int(break_input))
            except Exception:
                user_minutes = 0
            total_minutes = user_minutes + BASE_BREAK_MINUTES
            inst.休憩時間 = timedelta(minutes=total_minutes)

            # ✅ 把表单的 time/'HH:MM' 合成当天 datetime（带时区）存模型
            ci = form.cleaned_data.get("clock_in")
            co = form.cleaned_data.get("clock_out")
            unreturned = bool(form.cleaned_data.get("unreturned_flag"))

            ci_dt = _as_aware_dt(ci, report.date)
            co_dt = _as_aware_dt(co, report.date)

            if ci_dt is not None:
                inst.clock_in = ci_dt

            # 若已填写退勤时间，则视为已完成，覆盖勾选框
            if co_dt is not None:
                unreturned = False

            # 规则：
            # - 勾选“未完成入库手续” -> 退勤必须为空
            # - 未勾选：只有用户真的填了退勤才保存，否则保持为空
            if unreturned or co_dt is None:
                inst.clock_out = None
            else:
                inst.clock_out = co_dt

            try:
                inst.calculate_work_times()
            except Exception:
                pass

            inst.edited_by = request.user

            # ===== 保存主表/明细后，联动预约状态 =====
            #   - 退勤为空 + 勾选 -> status=未完成入库手续，actual_return 保持 None
            #   - 退勤有值 -> status=已完成（actual_return 会由 signals 用 inst.clock_out 同步回预约）
            try:
                from dailyreport.signals import _pick_reservation_for_report
                res = _pick_reservation_for_report(inst)
                if res:
                    if inst.clock_out is None:
                        # 退勤为空：实际入库也保持空
                        res.actual_return = None
                        if unreturned:
                            # 勾选“未完成入库手续”
                            try:
                                from vehicles.models import ReservationStatus
                                res.status = ReservationStatus.INCOMPLETE  # ← 使用新的枚举
                            except Exception:
                                res.status = "未完成出入库手续"
                            res.save(update_fields=["actual_return", "status"])
                        else:
                            res.save(update_fields=["actual_return"])
                    else:
                        # 退勤有值 => 已完成（actual_return 会由 signals 用 inst.clock_out 同步）
                        try:
                            res.status = ReservationStatus.DONE
                        except Exception:
                            # 兜底也用英文值，避免混入中文
                            res.status = "done"
                        res.save(update_fields=["status"])

                        # >>> BEGIN patch: finalize report times and status (views)
                        from django.utils import timezone
                        from dailyreport.models import DriverDailyReport

                        changed_fields_for_report = []

                        # 用本地时区把预约的实际出入库写回日报的 time 字段（避免 Time vs UTC DateTime 比较错误）
                        if getattr(res, "actual_departure", None):
                            _t_in = timezone.localtime(res.actual_departure).time()
                            if report.clock_in != _t_in:
                                report.clock_in = _t_in
                                changed_fields_for_report.append("clock_in")

                        if getattr(res, "actual_return", None):
                            _t_out = timezone.localtime(res.actual_return).time()
                            if report.clock_out != _t_out:
                                report.clock_out = _t_out
                                changed_fields_for_report.append("clock_out")

                        # 若日报已有出勤/退勤，则显式把状态置为已完成（completed）
                        if report.clock_in and report.clock_out and report.status != DriverDailyReport.STATUS_COMPLETED:
                            report.status = DriverDailyReport.STATUS_COMPLETED
                            changed_fields_for_report.append("status")

                        if changed_fields_for_report:
                            report.save(update_fields=changed_fields_for_report)
                        # >>> END patch

            except Exception as _e:
                logger.warning("update reservation status (incomplete/done) failed: %s", _e)
                
            # ===== [END] 预约状态联动 =====

            cash_total = sum(
                (it.cleaned_data.get('meter_fee') or 0)
                for it in formset.forms
                if it.cleaned_data.get('payment_method') == 'cash'
                and not it.cleaned_data.get('DELETE', False)
            )
            charter_cash_total = sum(
                (it.cleaned_data.get('charter_amount_jpy') or 0)
                for it in formset.forms
                if it.cleaned_data.get('is_charter')
                and (it.cleaned_data.get('charter_payment_method') in [
                    'jpy_cash', 'jp_cash', 'cash', 'rmb_cash',
                    'self_wechat', 'boss_wechat'
                ])
                and not it.cleaned_data.get('DELETE', False)
            )
            deposit = inst.deposit_amount or 0
            #inst.deposit_difference = deposit - cash_total - charter_cash_total
            inst = form.save(commit=False)

            inst.save()

            # === 明细行保存（强制 + 调试输出） ===
            formset.instance = inst

            # 调试：看一下管理表单和每一行的 cleaned_data
            try:
                print("DEBUG formset TOTAL_FORMS =", formset.total_form_count())
                print("DEBUG formset INITIAL_FORMS =", formset.initial_form_count())
                for idx, f in enumerate(formset.forms):
                    cd = getattr(f, "cleaned_data", None)
                    print(f"  [FORM {idx}] cleaned_data =", cd)
            except Exception as _e:
                print("DEBUG formset inspect failed:", _e)

            # 先拿到需要保存的对象列表（不含 DELETE 的）
            items = formset.save(commit=False)

            # 先处理删除的行，确保真的从数据库删掉
            for obj in formset.deleted_objects:
                try:
                    print("  [DELETE] item id =", obj.id)
                except Exception:
                    pass
                obj.delete()

            # 再保存新增/修改的行
            for item in items:
                # 防御：确保外键指向当前日报
                if getattr(item, "report_id", None) is None:
                    item.report = inst

                # 默认 is_pending=False（如果你需要这个行为）
                if getattr(item, "is_pending", None) is None:
                    item.is_pending = False

                item.save()
                try:
                    print(
                        "  [SAVE] item id =", item.id,
                        "meter_fee =", getattr(item, "meter_fee", None),
                        "is_charter =", getattr(item, "is_charter", None),
                        "payment_method =", getattr(item, "payment_method", None),
                        "charter_payment_method =", getattr(item, "charter_payment_method", None),
                    )
                except Exception:
                    pass

            try:
                # 保存完之后，再看一下这个日报下现在有多少条明细
                print("DEBUG after save -> inst.items.count() =", inst.items.count())
            except Exception as _e:
                print("DEBUG count items failed:", _e)

            # >>> [SYNC-RESERVATION CALL]
            try:
                _sync_reservation_actual_for_report(inst, _old_in, _old_out)
            except Exception as _e:
                logger.warning("sync reservation (dailyreport_edit_for_driver) failed: %s", _e)
            # <<< [SYNC-RESERVATION CALL]

            try:
                inst.has_issue = inst.items.filter(has_issue=True).exists()
                inst.save(update_fields=["has_issue"])
            except Exception:
                pass

            messages.success(request, "保存しました。")
            return redirect("dailyreport:driver_dailyreport_edit",
                            driver_id=driver.id, report_id=inst.id)
        else:
            messages.error(request, "❌ 保存失败，请检查输入内容")
    else:
        _prefill_report_without_fk(report)
        form = DriverDailyReportForm(instance=report)
        # 🔧 修复点：GET 也要用相同前缀，确保模板渲染的管理表单名称一致
        formset = ReportItemFormSet(instance=report, prefix=PREFIX)

    # ---------- 预填：尝试从 Reservation 带出车辆与实际出/入库（仅 GET，不写库） ----------
        try:
            from django.db.models import Q
            from vehicles.models import Reservation

            # 同一天的预约
            res_qs = Reservation.objects.filter(date=report.date)
            print("[prefill] report.id=", report.id, "report.date=", report.date)

            # 司机匹配：兼容“日报用档案ID、预约用账号ID”的场景
            d = report.driver
            user_obj = getattr(d, "user", None) or getattr(d, "account", None) \
                       or getattr(d, "auth_user", None) or getattr(d, "profile_user", None)
            cand = Q()
            if user_obj and getattr(user_obj, "id", None):
                cand |= Q(driver_id=user_obj.id)
                cand |= Q(driver__username=getattr(user_obj, "username", None))
            # 兜底：万一两边引用的是同一张表
            cand |= Q(driver_id=getattr(d, "id", None))
            res_qs = res_qs.filter(cand)

            # 若日报已选车，则进一步按车辆过滤
            if getattr(report, "vehicle_id", None):
                res_qs = res_qs.filter(vehicle_id=report.vehicle_id)

            # 优先选择“reserved/done”的预约；没有再取最早一条
            preferred = res_qs.filter(status__in=["reserved", "done"]).order_by("start_time").first()
            res = preferred or res_qs.order_by("start_time").first()
            print("[prefill] matched reservation ->",
                    None if not res else dict(
                        id=res.id,
                        vehicle_id=res.vehicle_id,
                        actual_departure=res.actual_departure,
                        actual_return=res.actual_return,
                        status=res.status,
                    ))


            if res:
                # 1) 预填车辆（日报未选车，预约有车）
                if not getattr(report, "vehicle_id", None) and getattr(res, "vehicle_id", None):
                    form.initial["vehicle"] = res.vehicle_id
                    if "vehicle" in form.fields:
                        form.fields["vehicle"].initial = res.vehicle_id  # 双保险：字段级 initial

                # 2) 预填出勤/退勤（日报为空，预约有“实际出/入库”），仅填 HH:MM
                if not getattr(report, "clock_in", None) and getattr(res, "actual_departure", None):
#                    form.initial["clock_in"] = res.actual_departure.astimezone().strftime("%H:%M") \
#                        if hasattr(res.actual_departure, "astimezone") else res.actual_departure.strftime("%H:%M")
                    _in = res.actual_departure
                    hhmm_in = (_in.astimezone().strftime("%H:%M") if hasattr(_in, "astimezone") else _in.strftime("%H:%M"))
                    form.initial["clock_in"] = hhmm_in
                    if "clock_in" in form.fields:
                        form.fields["clock_in"].initial = hhmm_in
                
                
                if not getattr(report, "clock_out", None) and getattr(res, "actual_return", None):
#                    form.initial["clock_out"] = res.actual_return.astimezone().strftime("%H:%M") \
#                        if hasattr(res.actual_return, "astimezone") else res.actual_return.strftime("%H:%M")
                    _out = res.actual_return
                    hhmm_out = (_out.astimezone().strftime("%H:%M") if hasattr(_out, "astimezone") else _out.strftime("%H:%M"))
                    form.initial["clock_out"] = hhmm_out
                    if "clock_out" in form.fields:
                        form.fields["clock_out"].initial = hhmm_out
        except Exception:
            # 静默容错：预填失败不影响页面打开
            pass

    data_iter = []
    for f in formset.forms:
        if f.is_bound and f.is_valid():
            cleaned = f.cleaned_data
            if not cleaned.get("DELETE", False):
                data_iter.append({
                    'meter_fee': _to_int0(cleaned.get('meter_fee')),
                    'payment_method': cleaned.get('payment_method') or '',
                    'note': cleaned.get('note') or '',
                    'DELETE': False,
                })
        elif f.instance and not getattr(f.instance, 'DELETE', False):
            data_iter.append({
                'meter_fee': _to_int0(getattr(f.instance, 'meter_fee', 0)),
                'payment_method': getattr(f.instance, 'payment_method', '') or '',
                'note': getattr(f.instance, 'note', '') or '',
                'DELETE': False,
            })
    totals_raw = calculate_totals_from_formset(data_iter)
    totals = {f"{k}_raw": v["total"] for k, v in totals_raw.items() if isinstance(v, dict)}
    totals.update({f"{k}_split": v["bonus"] for k, v in totals_raw.items() if isinstance(v, dict)})
    totals["meter_only_total"] = totals_raw.get("meter_only_total", 0)

    summary_keys = [
        ('meter', 'メーター(水揚)'),
        ('cash', '現金(ながし)'),
        ('uber', 'Uber'),
        ('didi', 'Didi'),
        ('credit', 'クレジ'),
        ('kyokushin', '京交信'),
        ('omron', 'オムロン'),
        ('kyotoshi', '京都市他'),
        ('qr', '扫码'),
    ]

    actual_break_min = _minutes_from_timedelta(getattr(report, "休憩時間", None))
    input_break_min  = max(0, actual_break_min - BASE_BREAK_MINUTES)
    break_time_h, break_time_m = divmod(input_break_min, 60)
    break_time_m = f"{break_time_m:02d}"
    actual_break_value = f"{actual_break_min // 60}:{actual_break_min % 60:02d}"

    # ✅ 统一口径的入金对比（ながし現金 + 貸切現金）
    deposit_summary = _build_deposit_summary_from_totals_raw(
        totals_raw,
        getattr(report, "deposit_amount", 0),
    )


    return render(request, 'dailyreport/driver_dailyreport_edit.html', {
        'form': form,
        'formset': formset,
        'driver': driver,
        'report': report,
        'is_edit': True,
        'summary_keys': summary_keys,
        'totals': totals,
        'break_time_h': break_time_h,
        'break_time_m': break_time_m,
        'actual_break_value': actual_break_value,
        'deposit_summary': deposit_summary,  # 👈 新增这一行
    })

# ========= 未分配账号司机：当天创建 =========
@user_passes_test(is_dailyreport_admin)
def driver_dailyreport_add_unassigned(request, driver_id):
    driver = get_object_or_404(Driver, id=driver_id, user__isnull=True)
    if not driver or driver.user:
        messages.warning(request, "未找到未分配账号的员工")
        return redirect("dailyreport:dailyreport_overview")

    today = date.today()
    report, created = DriverDailyReport.objects.get_or_create(
        driver=driver,
        date=today,
        defaults={"status": "草稿"}
    )
    print("🚗 创建日报：", driver.id, report.id, "是否新建：", created)

    if created:
        messages.success(request, f"已为 {driver.name} 创建 {today} 的日报。")
    else:
        messages.info(request, f"{driver.name} 今天的日报已存在，跳转到编辑页面。")

    return redirect("dailyreport:driver_dailyreport_edit", driver_id=driver.id, report_id=report.id)


# ===== [DISABLE LEGACY ENDPOINT BEGIN] =====
from django.http import Http404

@login_required
def my_dailyreports(request):
    """
    ⚠️ 旧入口已停用
    司机个人月度页请使用 vehicles.views.my_dailyreports ( /vehicles/my_dailyreports/ )
    """
    raise Http404("This endpoint is deprecated. Use /vehicles/my_dailyreports/.")
# ===== [DISABLE LEGACY ENDPOINT END] =====


# ========= 批量补账号 =========
@user_passes_test(is_dailyreport_admin)
def bind_missing_users(request):
    drivers_without_user = Driver.objects.filter(user__isnull=True)

    if request.method == 'POST':
        for driver in drivers_without_user:
            username = f"driver{driver.driver_code}"
            if not User.objects.filter(username=username).exists():
                user = User.objects.create_user(username=username, password='12345678')
                driver.user = user
                driver.save()
        return redirect('sdailyreport:bind_missing_users')

    return render(request, 'dailyreport/bind_missing_users.html', {
        'drivers': drivers_without_user,
    })

# ========= 导出：ETC 明细 =========
@user_passes_test(is_dailyreport_admin)
def export_etc_daily_csv(request, year, month):
    reports = DriverDailyReport.objects.filter(date__year=year, date__month=month)

    response = HttpResponse(content_type='text/csv')
    filename = f"ETC_日報明細_{year}-{month:02d}.csv"
    response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}"'

    writer = csv.writer(response)
    writer.writerow(['日期', '司机', 'ETC应收（円）', 'ETC实收（円）', '未收差额（円）'])

    for report in reports.order_by('date', 'driver__name'):
        expected = report.etc_expected or 0
        collected = report.etc_collected or 0
        diff = expected - collected

        writer.writerow([
            report.date.strftime('%Y-%m-%d'),
            report.driver.name if report.driver else "",
            expected,
            collected,
            diff
        ])

    return response

# ========= 导出：车辆运输实绩 =========
@user_passes_test(is_dailyreport_admin)
def export_vehicle_csv(request, year, month):
    reports = DriverDailyReport.objects.filter(
        date__year=year,
        date__month=month,
        vehicle__isnull=False
    ).select_related('vehicle')

    # ====== 车辆维度聚合容器 ======
    data = defaultdict(lambda: {
        '出勤日数': 0,
        '走行距離': 0,
        '実車距離': 0,
        '乗車回数': 0,

        '男性': 0,
        '女性': 0,
        '人数': 0,   # 男 + 女

        '水揚金額': 0,
        '車名': '',
        '車牌': '',
        '部門': '',
        '使用者名': '',
        '所有者名': '',
    })

    for r in reports:
        car = r.vehicle
        if not car:
            continue

        key = car.id
        mileage = float(r.mileage or 0)
        total_fee = float(r.total_meter_fee or 0)
        boarding_count = r.items.count()

        # ===== 出勤日数判定（基于当前模型字段）=====
        # 只要当天存在至少一条 ride_time 不为空的明细，即计 1 天
        if r.items.filter(ride_time__isnull=False).exists():
            data[key]['出勤日数'] += 1

        # ===== 距离 / 次数 =====
        data[key]['走行距離'] += mileage
        data[key]['実車距離'] += mileage * 0.75
        data[key]['乗車回数'] += boarding_count

        # ===== 真实乘客人数（男 / 女）=====
        items = r.items.all()
        agg = items.aggregate(
            male=Sum('num_male'),
            female=Sum('num_female'),
        )
        male = agg['male'] or 0
        female = agg['female'] or 0
        total_people = male + female

        data[key]['男性'] += male
        data[key]['女性'] += female
        data[key]['人数'] += total_people

        # ===== 金额 & 车辆信息 =====
        data[key]['水揚金額'] += total_fee
        data[key]['車名'] = car.name
        data[key]['車牌'] = car.license_plate
        data[key]['部門'] = getattr(car, 'department', '')
        data[key]['使用者名'] = getattr(car, 'user_company_name', '')
        data[key]['所有者名'] = getattr(car, 'owner_company_name', '')

    # ===== CSV 输出 =====
    response = HttpResponse(content_type='text/csv')
    filename = f"{year}年{month}月_車両運輸実績表.csv"
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"

    response.write(u'\ufeff'.encode('utf8'))  # BOM
    writer = csv.writer(response)

    # ====== 【PATCH】headers 加 人均売上 BEGIN ======
    headers = [
        '車名', '車牌', '部門', '使用者名', '所有者名',
        '出勤日数', '走行距離', '実車距離', '乗車回数',
        '男性', '女性', '人数',
        '平均每趟人数',
        '水揚金額',
        '人均売上',
    ]
    # ====== 【PATCH】END ======
    writer.writerow(headers)

    # ===== 行数据 + 合计 =====
    total = {
        '出勤日数': 0,
        '走行距離': 0,
        '実車距離': 0,
        '乗車回数': 0,
        '男性': 0,
        '女性': 0,
        '人数': 0,
        '水揚金額': 0,
    }

    for info in data.values():
        # ====== 【PATCH】平均每趟人数 计算 BEGIN ======
        avg_per_trip = (
            info['人数'] / info['乗車回数']
            if info['乗車回数'] > 0 else 0
        )
        # ====== 【PATCH】END ======
        # ====== 【PATCH】人均売上 计算 BEGIN ======
        sales_per_person = (
            info['水揚金額'] / info['人数']
            if info['人数'] > 0 else 0
        )
        # ====== 【PATCH】END ======

        row = [
            info['車名'],
            info['車牌'],
            info['部門'],
            info['使用者名'],
            info['所有者名'],
            info['出勤日数'],
            info['走行距離'],
            round(info['実車距離'], 2),
            info['乗車回数'],
            info['男性'],
            info['女性'],
            info['人数'],
            round(avg_per_trip, 2),
            round(info['水揚金額'], 2),
            round(sales_per_person, 2),
        ]
        writer.writerow(row)

        total['出勤日数'] += info['出勤日数']
        total['走行距離'] += info['走行距離']
        total['実車距離'] += info['実車距離']
        total['乗車回数'] += info['乗車回数']
        total['男性'] += info['男性']
        total['女性'] += info['女性']
        total['人数'] += info['人数']
        total['水揚金額'] += info['水揚金額']

    # ====== 【PATCH】合计 平均每趟人数 BEGIN ======
    total_avg = (
        total['人数'] / total['乗車回数']
        if total['乗車回数'] > 0 else 0
    )
    # ====== 【PATCH】合计 人均売上 BEGIN ======
    total_sales_per_person = (
        total['水揚金額'] / total['人数']
        if total['人数'] > 0 else 0
    )
    # ====== 【PATCH】END ======

    writer.writerow([
        '合計', '', '', '', '',
        total['出勤日数'],
        total['走行距離'],
        round(total['実車距離'], 2),
        total['乗車回数'],
        total['男性'],
        total['女性'],
        total['人数'],
        round(total_avg, 2),
        round(total['水揚金額'], 2),
        round(total_sales_per_person, 2),
    ])
    # ====== 【PATCH】END ======

    return response

# ========= 月份入口（表单选择） =========
@user_passes_test(is_dailyreport_admin)
def dailyreport_add_by_month(request, driver_id):
    driver = get_object_or_404(Driver, pk=driver_id)

    month_str = request.GET.get("month")
    if not month_str:
        return redirect("dailyreport:driver_dailyreport_add_selector", driver_id=driver_id)

    try:
        year, month = map(int, month_str.split("-"))
        assert 1 <= month <= 12
    except (ValueError, AssertionError):
        return redirect("dailyreport:driver_dailyreport_add_selector", driver_id=driver_id)

    current_month = f"{year}年{month}月"

    if request.method == "POST":
        selected_date_str = request.POST.get("selected_date")
        try:
            selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return render(request, "dailyreport/driver_dailyreport_add.html", {
                "driver": driver, "year": year, "month": month,
                "current_month": current_month, "error": "日付が正しくありません"
            })

        base_url = reverse("dailyreport:driver_dailyreport_direct_add", args=[driver.id])
        query_string = urlencode({"date": selected_date})
        url = f"{base_url}?{query_string}"
        return redirect(url)

    return render(request, "dailyreport/driver_dailyreport_add.html", {
        "driver": driver,
        "year": year,
        "month": month,
        "current_month": current_month,
    })

# ========= 月度总览 =========
@user_passes_test(is_dailyreport_admin)
def dailyreport_overview(request):
    today = now().date()
    keyword = request.GET.get('keyword', '').strip()
    month_str = request.GET.get('month', '')

    try:
        month = datetime.strptime(month_str, "%Y-%m")
    except ValueError:
        month = today.replace(day=1)
        month_str = month.strftime('%Y-%m')

    month_label = f"{month.year}年{month.month:02d}月"
    prev_month = (month - relativedelta(months=1)).strftime('%Y-%m')
    next_month = (month + relativedelta(months=1)).strftime('%Y-%m')

    export_year = month.year
    export_month = month.month

    # ========== [BEGIN 保留：原来的按业务日期月份过滤] ==========
    # reports_all = DriverDailyReport.objects.filter(
    #     date__year=month.year,
    #     date__month=month.month,
    # )
    # ========== [END   保留：原来的按业务日期月份过滤] ==========

    # ========== [BEGIN 新：按“勤務開始日(开始日)”归属月份] ==========

    # 约定：clock_in < 06:00 视为夜勤跨零点 → 归前一天
    work_date_expr = Case(
        When(clock_in__lt=time(6, 0),
            then=ExpressionWrapper(F('date') - timedelta(days=1), output_field=DateField())),
        default=F('date'),
        output_field=DateField()
    )

    reports_all = (
        DriverDailyReport.objects
        .annotate(work_date=work_date_expr)
        .filter(work_date__year=month.year, work_date__month=month.month)
    )
    # ========== [END   新：按“勤務開始日(开始日)”归属月份] ==========

    drivers = get_active_drivers(month, keyword)

    if keyword:
        drivers = drivers.filter(
            Q(name__icontains=keyword) |
            Q(kana__icontains=keyword) |
            Q(driver_code__icontains=keyword)
        )

    reports = reports_all.filter(driver__in=drivers)

    # ========== [BEGIN 保留：旧写法，统计了所有司机（含已离职）] ==========
    # items_all = DriverDailyReportItem.objects.filter(report__in=reports_all)
    # ========== [END   保留] ==========

    # ========== [BEGIN 新写法：仅统计页面显示的司机（活跃/筛选后）] ==========
    items_all = DriverDailyReportItem.objects.filter(report__in=reports)
    # ========== [END   新写法] ==========
    items_norm = items_all.annotate(
        pm=Lower(Trim('payment_method')),
        cpm=Lower(Trim('charter_payment_method')),
    )

    totals = defaultdict(Decimal)
    counts = defaultdict(int)

    # === Uber 派生：予約 / チップ / プロモーション（严格匹配版）===

    # 显式字段别名（你上面已对 items_norm 做了 Lower/Trim）
    UBER_RESV_ALIASES  = {'uber_reservation', 'uber_resv', 'uber予約'}
    UBER_TIP_ALIASES   = {'uber_tip', 'uber tip', 'ubertip'}
    UBER_PROMO_ALIASES = {'uber_promo', 'uber_promotion', 'uberプロモーション'}

    # 仅精确匹配 payment_method/charter_payment_method；不再使用 note/comment 关键词
    _q_resv  = Q(pm__in=UBER_RESV_ALIASES)  | Q(cpm__in=UBER_RESV_ALIASES)
    _q_tip   = Q(pm__in=UBER_TIP_ALIASES)   | Q(cpm__in=UBER_TIP_ALIASES)
    _q_promo = Q(pm__in=UBER_PROMO_ALIASES) | Q(cpm__in=UBER_PROMO_ALIASES)

    _qs_resv  = items_norm.filter(_q_resv)
    _qs_tip   = items_norm.filter(_q_tip)
    _qs_promo = items_norm.filter(_q_promo)

    def _sum_amount_by_is_charter(qs):
        non_charter = qs.filter(is_charter=False).aggregate(x=Sum('meter_fee'))['x'] or Decimal('0')
        charter     = qs.filter(is_charter=True ).aggregate(x=Sum('charter_amount_jpy'))['x'] or Decimal('0')
        return non_charter + charter

    totals['uber_reservation_total'] = _sum_amount_by_is_charter(_qs_resv)
    totals['uber_tip_total']         = _sum_amount_by_is_charter(_qs_tip)
    totals['uber_promotion_total']   = _sum_amount_by_is_charter(_qs_promo)

    counts['uber_reservation'] = _qs_resv.count()
    counts['uber_tip']         = _qs_tip.count()
    counts['uber_promotion']   = _qs_promo.count()

    meter_sum_non_charter = items_norm.filter(is_charter=False)\
        .aggregate(x=Sum('meter_fee'))['x'] or Decimal('0')
    totals['total_meter'] = meter_sum_non_charter
    totals['meter_only_total'] = meter_sum_non_charter

    ALIASES = {
        'cash':      {'normal': ['cash'],                 'charter': ['jpy_cash']},
        'credit':    {'normal': ['credit', 'credit_card'],'charter': ['credit','credit_card']},
        'uber':      {'normal': ['uber'],                 'charter': ['uber']},
        'didi':      {'normal': ['didi'],                 'charter': ['didi']},
        'kyokushin': {'normal': ['kyokushin'],            'charter': ['kyokushin']},
        'omron':     {'normal': ['omron'],                'charter': ['omron']},
        'kyotoshi':  {'normal': ['kyotoshi'],             'charter': ['kyotoshi']},
        'qr':        {'normal': ['qr', 'scanpay'],        'charter': ['qr', 'scanpay']},
    }
    EXCLUDE_CHARTER_IN_METHODS = {'cash'}

    for key, alias in ALIASES.items():
        normal_qs = items_norm.filter(is_charter=False, pm__in=alias['normal'])
        normal_amt = normal_qs.aggregate(x=Sum('meter_fee'))['x'] or Decimal('0')
        normal_cnt = normal_qs.count()

        charter_amt = Decimal('0')
        charter_cnt = 0
        if key not in EXCLUDE_CHARTER_IN_METHODS:
            charter_qs = items_norm.filter(is_charter=True, cpm__in=alias['charter'])
            charter_amt = charter_qs.aggregate(x=Sum('charter_amount_jpy'))['x'] or Decimal('0')
            charter_cnt = charter_qs.count()

        totals[f'total_{key}'] = normal_amt + charter_amt
        counts[key] = normal_cnt + charter_cnt

    totals['charter_cash_total'] = items_norm.filter(
        is_charter=True, cpm__in=['jpy_cash']
    ).aggregate(x=Sum('charter_amount_jpy'))['x'] or Decimal('0')

    totals['charter_uncollected_total'] = items_norm.filter(
        is_charter=True, cpm__in=['to_company', 'invoice', 'uncollected', '未収', '請求']
    ).aggregate(x=Sum('charter_amount_jpy'))['x'] or Decimal('0')

    totals['total_meter'] = (
        (totals.get('meter_only_total') or Decimal('0')) +
        (totals.get('charter_cash_total') or Decimal('0')) +
        (totals.get('charter_uncollected_total') or Decimal('0'))
    )

    rates = {
        'meter':     Decimal('0.9091'),
        'cash':      Decimal('0'),
        'uber':      Decimal('0.05'),
        'didi':      Decimal('0.05'),
        'credit':    Decimal('0.05'),
        'kyokushin': Decimal('0.05'),
        'omron':     Decimal('0.05'),
        'kyotoshi':  Decimal('0.05'),
        'qr':        Decimal('0.05'),
    }

    def split(key):
        amt = totals.get(f"total_{key}") or Decimal('0')
        return (amt * rates[key]).quantize(Decimal('1'), rounding=ROUND_HALF_UP)

    totals_all = {k: {"total": totals.get(f"total_{k}", Decimal("0")), "bonus": split(k)} for k in rates}
    totals_all["meter_only_total"] = totals.get("meter_only_total", Decimal("0"))

    gross = totals.get('total_meter') or Decimal('0')
    totals['meter_pre_tax'] = (gross / Decimal('1.1')).quantize(Decimal('1'), rounding=ROUND_HALF_UP)

    etc_shortage_total = reports.aggregate(total=Sum('etc_shortage'))['total'] or 0

    items = DriverDailyReportItem.objects.filter(report__in=reports)
    per_driver = items.values('report__driver').annotate(
        meter_only=Sum('meter_fee', filter=Q(is_charter=False)),
        charter_cash=Sum(
            'charter_amount_jpy',
            filter=Q(is_charter=True, charter_payment_method__in=['jpy_cash', 'jp_cash', 'cash'])
        ),
        charter_uncol=Sum(
            'charter_amount_jpy',
            filter=Q(is_charter=True, charter_payment_method__in=['to_company', 'invoice', 'uncollected', '未収', '請求'])
        ),
    )

    fee_map = {
        r['report__driver']: (r['meter_only'] or 0) + (r['charter_cash'] or 0) + (r['charter_uncol'] or 0)
        for r in per_driver
    }

    sort = request.GET.get("sort", "amount_desc")

    def code_key(d):
        code = (getattr(d, "driver_code", "") or "").strip()
        if code.isdigit():
            return (0, int(code))
        return (1, code)

    if sort == "code_asc":
        ordered_drivers = sorted(drivers, key=code_key)
    elif sort == "code_desc":
        ordered_drivers = sorted(drivers, key=code_key, reverse=True)
    elif sort == "amount_asc":
        ordered_drivers = sorted(
            drivers,
            key=lambda d: (fee_map.get(d.id, Decimal("0")), code_key(d))
        )
    else:
        ordered_drivers = sorted(
            drivers,
            key=lambda d: (fee_map.get(d.id, Decimal("0")), code_key(d)),
            reverse=True
        )

    driver_data = []
    for d in ordered_drivers:
        total = fee_map.get(d.id, Decimal("0"))
        has_any = d.id in fee_map
        has_issue = reports.filter(driver=d, has_issue=True).exists()
        note = "⚠️ 異常あり" if has_issue else ("（未報告）" if not has_any else "")
        driver_data.append({
            'driver': d,
            'total_fee': total,
            'note': note,
            'month_str': month_str,
        })

    page_obj = Paginator(driver_data, 10).get_page(request.GET.get('page'))

    summary_keys = [
        ('meter', 'メーター(水揚)'),
        ('cash', '現金'),
        ('uber', 'Uber'),
        ('didi', 'Didi'),
        ('credit', 'クレジットカード'),
        ('kyokushin', '京交信'),
        ('omron', 'オムロン'),
        ('kyotoshi', '京都市他'),
        ('qr', '扫码'),
    ]

    return render(request, 'dailyreport/dailyreport_overview.html', {
        'totals': totals,
        'totals_all': totals_all,
        'etc_shortage_total': etc_shortage_total,
        'drivers': drivers,
        'page_obj': page_obj,
        'counts': counts,
        'current_sort': sort,
        'keyword': keyword,
        'month_str': month_str,
        'current_year': export_year,
        'current_month': export_month,
        'summary_keys': summary_keys,
        'month_label': month_label,
        'prev_month': prev_month,
        'next_month': next_month,
        'sort': sort,
    })


# ===== BEGIN IMPORT_EXTERNAL_DAILYREPORT_HELPERS M3 =====
def _to_bool(val):
    """
    Excelセルの 1/0/TRUE/FALSE/空 を Python bool に変換
    """
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("1", "true", "yes", "y", "はい", "t"):
        return True
    return False


# ===== BEGIN REPLACE DUP-HELPER: 外部日報Excelの重複チェック（時間を正規化） =====
def find_duplicate_rows_in_external_excel(file_bytes: bytes):
    """
    外部日報 Excel 内の「同一ドライバー＋同一日付」で
    『時間／乗車地／降車地』が完全一致している行を検出する。

    ・時間は "9:00" / "09:00" / "09:00:00" / Excel 時刻セル などを
      すべて "HH:MM" 形式に正規化してから比較する。
    """

    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    # シート選択：DailyReport があれば優先、なければアクティブ
    if "DailyReport" in wb.sheetnames:
        ws = wb["DailyReport"]
    else:
        ws = wb.active

    # ヘッダ行
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return []

    col_index = {}
    for idx, name in enumerate(header_row):
        if not name:
            continue
        col_index[str(name).strip().lower()] = idx

    # 重複判定に必要なカラム
    required_for_dup = ["date", "driver_code", "ride_time", "ride_from", "ride_to"]
    missing = [c for c in required_for_dup if c not in col_index]
    if missing:
        # テンプレートが古い／壊れている → 重複チェックは諦めて通常処理に任せる
        return []

    def get(row, key, default=None):
        idx = col_index.get(key)
        if idx is None:
            return default
        return row[idx]

    def _parse_date(v):
        if v in (None, ""):
            return None
        if isinstance(v, (datetime, date)):
            return v.date() if isinstance(v, datetime) else v
        try:
            return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()
        except Exception:
            return None

    def _normalize_time_str(v):
        """
        時刻セルを 'HH:MM' に統一する。
        - Excel の時刻セル（datetime.time / datetime） → strftime('%H:%M')
        - 文字列 '9:00' / '09:00:00' など → できる限りパースして HH:MM にする
        """
        if v in (None, ""):
            return ""

        # datetime / time の場合
        if isinstance(v, datetime):
            return v.time().strftime("%H:%M")
        if isinstance(v, time):
            return v.strftime("%H:%M")

        s = str(v).strip()
        if not s:
            return ""

        # 既に HH:MM or H:MM っぽい場合の簡易処理
        if ":" in s:
            parts = s.split(":")
            try:
                h = int(parts[0])
                m = int(parts[1])
                return f"{h:02d}:{m:02d}"
            except Exception:
                pass

        # 数値（例えば 0.375 = 9:00）で来た場合は、24時間をかけて変換してみる
        try:
            num = float(s)
            total_minutes = int(round(num * 24 * 60))
            h = (total_minutes // 60) % 24
            m = total_minutes % 60
            return f"{h:02d}:{m:02d}"
        except Exception:
            # 最後の手段：そのまま返す（これでも完全一致なら同一とみなせる）
            return s

    duplicates = []
    seen = {}  # (driver_code, date_str, time_norm, from, to) -> first_row_idx

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        date_val = _parse_date(get(row, "date"))
        driver_code_raw = get(row, "driver_code")
        ride_time_raw = get(row, "ride_time")
        ride_from_val = get(row, "ride_from")
        ride_to_val = get(row, "ride_to")

        if not (date_val and driver_code_raw and ride_time_raw):
            continue  # キー要素が欠けていたらスキップ

        date_str = date_val.isoformat()
        driver_code = str(driver_code_raw).strip()
        time_norm = _normalize_time_str(ride_time_raw)
        from_str = str(ride_from_val or "").strip()
        to_str = str(ride_to_val or "").strip()

        key = (driver_code, date_str, time_norm, from_str, to_str)
        if key in seen:
            duplicates.append({
                "row": row_idx,
                "first_row": seen[key],
                "driver_code": driver_code,
                "date": date_str,
                "ride_time": time_norm,
                "ride_from": from_str,
                "ride_to": to_str,
            })
        else:
            seen[key] = row_idx

    return duplicates
# ===== END REPLACE DUP-HELPER =====



# ===== BEGIN REPLACE M-V2: 解析外部日報 Excel（增强模板版） =====
def parse_external_dailyreport_excel(uploaded_file, current_user=None):
    """
    解析由外部录入员填写的 Excel（增强模板版），
    一次性创建 / 更新 DriverDailyReport + DriverDailyReportItem。

    期望的 Sheet:
        - DailyReport: 主数据
        - MasterData: 支払コード列表（仅用于下拉，解析时只用 code）

    期望的表头（daily sheet 第 1 行）：
        date, driver_code, vehicle_number,
        clock_in, clock_out, break_time,
        gas_volume, mileage,
        ride_time, ride_from, ride_to,
        meter_fee, payment_method,
        is_charter, charter_amount, charter_payment_method,
        note, is_pending
    """
    wb = load_workbook(uploaded_file, data_only=True)

    # 1) 读取版本信息（如果 Excel 内部写了就校验，否则只做 header 检查）
    version_in_file = None
    if "Meta" in wb.sheetnames:
        meta_ws = wb["Meta"]
        raw = meta_ws["A1"].value if meta_ws["A1"].value else ""
        if isinstance(raw, str) and "TEMPLATE_VERSION" in raw:
            # 例： "TEMPLATE_VERSION=2025.01"
            try:
                version_in_file = raw.split("=", 1)[1].strip()
            except Exception:
                version_in_file = None

    if version_in_file and version_in_file != EXPECTED_TEMPLATE_VERSION:
        return {
            "ok": False,
            "version_in_file": version_in_file,
            "errors": [
                f"テンプレートのバージョンが違います。期待: {EXPECTED_TEMPLATE_VERSION}, ファイル内: {version_in_file}"
            ],
            "created_reports": 0,
            "updated_reports": 0,
            "created_items": 0,
        }

    # 2) 选择主 sheet：优先 DailyReport，其次 active
    if "DailyReport" in wb.sheetnames:
        ws = wb["DailyReport"]
    else:
        ws = wb.active

    # 3) 读取表头并构建列索引 map
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {}  # "date" -> 0, ...
    for idx, name in enumerate(header_row):
        if not name:
            continue
        key = str(name).strip().lower()
        col_index[key] = idx

    # 必须的列
    required_cols = [
        "date",
        "driver_code",
        "vehicle_number",
        "ride_time",
        "meter_fee",
        "payment_method",
    ]
    missing = [c for c in required_cols if c not in col_index]
    if missing:
        return {
            "ok": False,
            "version_in_file": version_in_file,
            "errors": [f"必須列が足りません: {', '.join(missing)}"],
            "created_reports": 0,
            "updated_reports": 0,
            "created_items": 0,
        }

    # 4) 工具函数
    def get(row, key, default=None):
        idx = col_index.get(key)
        if idx is None:
            return default
        return row[idx]

    def parse_date(v):
        if v in (None, ""):
            return None
        if isinstance(v, (datetime, date)):
            return v.date() if isinstance(v, datetime) else v
        try:
            return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()
        except Exception:
            return None

    def parse_time_cell(v):
        if v in (None, ""):
            return None
        if isinstance(v, datetime):
            return v.time()
        if isinstance(v, time):
            return v
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return None
            try:
                h, m = s.split(":")
                return time(hour=int(h), minute=int(m))
            except Exception:
                return None
        return None

    def parse_timedelta_hm(v):
        if v in (None, ""):
            return None
        if isinstance(v, timedelta):
            return v
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return None
            try:
                h, m = s.split(":")
                return timedelta(hours=int(h), minutes=int(m))
            except Exception:
                return None
        return None

    def parse_decimal(v):
        if v in (None, ""):
            return None
        try:
            return Decimal(str(v))
        except Exception:
            return None

    def parse_bool(v):
        if v in (None, ""):
            return False
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        return s in ("1", "true", "yes", "y", "t", "○", "◯")

    # 5) 逐行解析
    created_reports = 0
    updated_reports = 0
    created_items = 0
    errors = []

    # 为了避免重复 get_or_create，每个 (driver_id, date) 缓存一份 report
    report_cache = {}

    @transaction.atomic
    def _inner():
        nonlocal created_reports, updated_reports, created_items

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过全空行
            if not any(row):
                continue

            date_val = parse_date(get(row, "date"))
            driver_code_raw = get(row, "driver_code")
            vehicle_number = get(row, "vehicle_number")

            # ====== driver_code = 员工编号（Driver.driver_code） ======
            if not date_val or not driver_code_raw:
                errors.append(f"{row_idx} 行目: date または driver_code（员工编号）が空です。")
                continue

            driver_code_str = str(driver_code_raw).strip()

            # 司机：用 Driver.driver_code 匹配，而不是 pk
            try:
                driver = Driver.objects.get(driver_code=driver_code_str)
            except Driver.DoesNotExist:
                errors.append(
                    f"{row_idx} 行目: driver_code='{driver_code_str}'（员工编号）に該当するドライバーが見つかりません。"
                )
                continue
            except Driver.MultipleObjectsReturned:
                errors.append(
                    f"{row_idx} 行目: driver_code='{driver_code_str}'（员工编号）に該当するドライバーが複数存在します。"
                )
                continue
            # ====== driver_code 处理到此结束 ======

            # 车辆：车号匹配 license_plate 中包含 vehicle_number（例：xxx-5001）
            car = None
            if vehicle_number not in (None, ""):
                vn_str = str(vehicle_number).strip()
                car = Car.objects.filter(license_plate__contains=vn_str).first()
                if not car:
                    errors.append(f"{row_idx} 行目: vehicle_number={vn_str} に対応する車両が見つかりません。")
                    # 不致命：允许无车继续，只是日報 vehicle 为 None

            # 该司机+日期 对应的 key
            cache_key = (driver.pk, date_val)
            report = report_cache.get(cache_key)

            # 构造当行的“日報级”字段
            clock_in = parse_time_cell(get(row, "clock_in"))
            clock_out = parse_time_cell(get(row, "clock_out"))
            break_td = parse_timedelta_hm(get(row, "break_time"))
            gas_vol = parse_decimal(get(row, "gas_volume"))
            mileage = parse_decimal(get(row, "mileage"))

            if report is None:
                defaults = {
                    "vehicle": car,
                }
                if clock_in:
                    defaults["clock_in"] = clock_in
                if clock_out:
                    defaults["clock_out"] = clock_out
                if break_td is not None:
                    defaults["休憩時間"] = break_td
                if gas_vol is not None:
                    defaults["gas_volume"] = gas_vol
                if mileage is not None:
                    defaults["mileage"] = mileage

                report, created = DriverDailyReport.objects.get_or_create(
                    driver=driver,
                    date=date_val,
                    defaults=defaults,
                )
                if created:
                    created_reports += 1
                else:
                    # 已存在的情况，后面如果行里有非空值再更新
                    pass
                report_cache[cache_key] = report
            else:
                created = False

            changed = False
            if not created:
                if car and report.vehicle != car:
                    report.vehicle = car
                    changed = True
                if clock_in and report.clock_in != clock_in:
                    report.clock_in = clock_in
                    changed = True
                if clock_out and report.clock_out != clock_out:
                    report.clock_out = clock_out
                    changed = True
                if break_td is not None and report.休憩時間 != break_td:
                    report.休憩時間 = break_td
                    changed = True
                if gas_vol is not None and report.gas_volume != gas_vol:
                    report.gas_volume = gas_vol
                    changed = True
                if mileage is not None and report.mileage != mileage:
                    report.mileage = mileage
                    changed = True

            # 有出勤/退勤/休憩的任一信息时，重算勤務時間/実働/残業
            if clock_in or clock_out or break_td is not None:
                report.calculate_work_times()
                changed = True

            # 记录编辑人
            if current_user and getattr(report, "edited_by_id", None) != current_user.id:
                report.edited_by = current_user
                changed = True

            if changed:
                report.save()
                if not created:
                    updated_reports += 1

            # ====== 创建行明细 DriverDailyReportItem ======
            ride_time = get(row, "ride_time")
            ride_from = get(row, "ride_from")
            ride_to = get(row, "ride_to")
            meter_fee_val = parse_decimal(get(row, "meter_fee")) or Decimal("0")
            payment_method = (get(row, "payment_method") or "").strip() or None
            is_charter = parse_bool(get(row, "is_charter"))
            charter_amount = parse_decimal(get(row, "charter_amount"))
            charter_payment_method = (get(row, "charter_payment_method") or "").strip() or None
            note = get(row, "note") or ""
            is_pending = parse_bool(get(row, "is_pending"))

            # 如果行明细关键字段都空，则不创建 item
            if not (ride_time or ride_from or ride_to or meter_fee_val or payment_method):
                continue

            item = DriverDailyReportItem.objects.create(
                report=report,
                ride_time=str(ride_time or "").strip(),
                ride_from=str(ride_from or "").strip(),
                ride_to=str(ride_to or "").strip(),
                meter_fee=meter_fee_val,
                payment_method=payment_method or "",
                is_charter=is_charter,
                charter_amount_jpy=charter_amount,
                charter_payment_method=charter_payment_method,
                note=str(note or "").strip(),
                is_pending=is_pending,
            )
            created_items += 1

    # 原子操作，出错就整批回滚
    try:
        _inner()
    except Exception as e:
        errors.append(f"予期しないエラー: {e}")
        return {
            "ok": False,
            "version_in_file": version_in_file,
            "errors": errors,
            "created_reports": 0,
            "updated_reports": 0,
            "created_items": 0,
        }

    return {
        "ok": len(errors) == 0,
        "version_in_file": version_in_file,
        "errors": errors,
        "created_reports": created_reports,
        "updated_reports": updated_reports,
        "created_items": created_items,
    }
# ===== END REPLACE M-V2 =====

# ===== BEGIN IMPORT_EXTERNAL_DAILYREPORT_VIEW M5 (with duplicate check) =====
@login_required
@require_http_methods(["GET", "POST"])
def external_dailyreport_import(request):
    """
    外部录入员做好的 Excel を取り込む画面＋処理。

    フロー：
      1) 通常 POST（file 付き）:
         - Excel をバイト列に読み込み
         - find_duplicate_rows_in_external_excel() で
           「同一ドライバー＋同一日付＋時間＋乗車地＋降車地」の重複行を検出
         - 重複なし → そのまま parse_external_dailyreport_excel() で取り込み
         - 重複あり → DB には書き込まず、重複一覧＋base64化したファイルを
           external_import.html に渡して「確認画面」を表示

      2) 確認画面からの POST（confirm_duplicates=1, file_base64付き）:
         - base64 から元の Excel バイト列を復元し、
           parse_external_dailyreport_excel() を実行して実際に取り込み。
    """
    # ---- ② 確認後の再取込（file_base64 経由） ----
    if (
        request.method == "POST"
        and request.POST.get("confirm_duplicates") == "1"
        and "file" not in request.FILES
    ):
        b64 = request.POST.get("file_base64", "")
        if not b64:
            messages.error(request, "重複確認後の再取込に失敗しました（ファイル情報が見つかりません）。もう一度ファイルを選択してください。")
            form = ExternalDailyReportImportForm()
            return render(
                request,
                "dailyreport/external_import.html",
                {
                    "form": form,
                    "template_version": TEMPLATE_VERSION,
                    "import_result": None,
                    "duplicate_warnings": None,
                    "file_base64": "",
                },
            )

        try:
            file_bytes = base64.b64decode(b64)
        except Exception:
            messages.error(request, "ファイル情報の復元に失敗しました。もう一度ファイルを選択してください。")
            form = ExternalDailyReportImportForm()
            return render(
                request,
                "dailyreport/external_import.html",
                {
                    "form": form,
                    "template_version": TEMPLATE_VERSION,
                    "import_result": None,
                    "duplicate_warnings": None,
                    "file_base64": "",
                },
            )

        # ここで本番取り込み
        result = parse_external_dailyreport_excel(BytesIO(file_bytes), current_user=request.user)

        if result["ok"]:
            messages.success(
                request,
                f"取込完了：日報 {result['created_reports']} 件新規 / "
                f"{result['updated_reports']} 件更新、明細 {result['created_items']} 行。"
            )
        else:
            messages.error(request, "一部エラーがあります。内容を確認してください。")

        return render(
            request,
            "dailyreport/external_import.html",
            {
                "form": ExternalDailyReportImportForm(),
                "template_version": TEMPLATE_VERSION,
                "import_result": result,
                "duplicate_warnings": None,
                "file_base64": "",
            },
        )

    # ---- ① 通常のアップロード（最初の取り込みボタン） ----
    if request.method == "POST":
        form = ExternalDailyReportImportForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data["file"]

            # アップロードファイルを丸ごと bytes に読み込む
            file_bytes = uploaded_file.read()

            # まず重複チェックだけ行う（DB には書き込まない）
            duplicate_warnings = find_duplicate_rows_in_external_excel(file_bytes)

            if duplicate_warnings:
                # 重複あり → ここでは取り込まず、確認画面を表示
                file_b64 = base64.b64encode(file_bytes).decode("ascii")
                messages.warning(
                    request,
                    "同じ日付・ドライバーで『時間／乗車地／降車地』が重複している明細が検出されました。内容を確認してください。"
                )
                return render(
                    request,
                    "dailyreport/external_import.html",
                    {
                        "form": ExternalDailyReportImportForm(),  # 新しいファイルを選び直すこともできる
                        "template_version": TEMPLATE_VERSION,
                        "import_result": None,
                        "duplicate_warnings": duplicate_warnings,
                        "file_base64": file_b64,
                    },
                )

            # 重複なし → そのまま取り込む
            result = parse_external_dailyreport_excel(BytesIO(file_bytes), current_user=request.user)

            if result["ok"]:
                messages.success(
                    request,
                    f"取込完了：日報 {result['created_reports']} 件新規 / "
                    f"{result['updated_reports']} 件更新、明細 {result['created_items']} 行。"
                )
            else:
                messages.error(request, "一部エラーがあります。内容を確認してください。")

            return render(
                request,
                "dailyreport/external_import.html",
                {
                    "form": ExternalDailyReportImportForm(),
                    "template_version": TEMPLATE_VERSION,
                    "import_result": result,
                    "duplicate_warnings": None,
                    "file_base64": "",
                },
            )

    # ---- GET or フォームエラー時 ----
    else:
        form = ExternalDailyReportImportForm()

    return render(
        request,
        "dailyreport/external_import.html",
        {
            "form": form,
            "template_version": TEMPLATE_VERSION,
            "import_result": None,
            "duplicate_warnings": None,
            "file_base64": "",
        },
    )
# ===== END IMPORT_EXTERNAL_DAILYREPORT_VIEW M5 =====